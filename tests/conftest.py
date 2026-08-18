from __future__ import annotations

import os
import sys
from datetime import date
from pathlib import Path

import pytest
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from processor import SUMMARY_LABELS, build_target_day_columns

os.environ.setdefault("QT_QPA_PLATFORM", "offscreen")


@pytest.fixture(scope="session")
def qapp():
    from PySide6.QtWidgets import QApplication

    return QApplication.instance() or QApplication(sys.argv)


def build_source_workbook(path: Path, zero_days: set[str] | None = None) -> None:
    zero_days = zero_days or set()
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    ws["A1"] = "Aprilie 2026"
    ws.append(["Data", "Bov", "Nr.buc", "Ov", "Nr.buc", "Nr.tranșatori", "Nr.fasonatori", "Observatii"])
    source_days = [
        ("01.04.2026", 1000, 120, None, None, ["16T"]),
        ("02.04.2026", 900, 100, None, 12, ["14T(2reprize)", "16T(1repriza)"]),
        ("03.04.2026", 920, 110, None, None, ["15T"]),
        ("04.04.2026", 950, 130, 6, None, ["12T(1repriza)", "18T(2reprize)"]),
        ("06.04.2026", 1100, 150, None, 6, ["15T"]),
    ]

    for row_date, bovine_total, bovine_count, ovine_primary, ovine_fallback, transatori_entries in source_days:
        if row_date in zero_days:
            ws.append([row_date, 0, 0, None, None, None, "17F", None])
            continue

        ws.append(
            [
                row_date,
                bovine_total,
                bovine_count,
                ovine_primary,
                ovine_fallback,
                transatori_entries[0],
                "17F",
                None,
            ]
        )
        for transatori_entry in transatori_entries[1:]:
            ws.append([None, None, None, None, None, transatori_entry, None, None])

    ws.append(["Total", 3770, None, None, None, None, None, None])

    wb.save(path)
    wb.close()


def build_source_workbook_for_dates(path: Path, source_dates: list[date]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Aprilie 2026"
    ws.append(["Data", "Bov", "Nr.buc", "Ov", "Nr.buc", "Nr.tranșatori", "Nr.fasonatori", "Observatii"])

    for index, source_date in enumerate(source_dates, start=1):
        ws.append(
            [
                source_date.strftime("%d.%m.%Y"),
                1000 + index,
                100 + index,
                None,
                None,
                "15T",
                "17F",
                None,
            ]
        )

    ws.append(["Total", None, None, None, None, None, None, None])
    wb.save(path)
    wb.close()


def build_source_pdf(path: Path) -> None:
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24)
    styles = getSampleStyleSheet()
    data = [
        ["Data", "Bov", "Nr.buc", "Ov", "Nr.buc", "Nr.transatori", "Nr.fasonatori", "Observatii"],
        ["01.04.2026", "1000", "120", "", "", "16T", "17F", ""],
        ["02.04.2026", "900", "100", "", "12", "14T(2reprize)", "17F", ""],
        ["", "", "", "", "", "16T(1repriza)", "", ""],
        ["03.04.2026", "920", "110", "", "", "15T", "17F", ""],
        ["04.04.2026", "950", "130", "6", "", "12T(1repriza)", "17F", ""],
        ["", "", "", "", "", "18T(2reprize)", "", ""],
        ["06.04.2026", "1100", "150", "", "6", "15T", "17F", ""],
        ["Total", "3770", "", "", "", "", "", ""],
    ]
    table = Table(data, colWidths=[72, 54, 54, 44, 54, 112, 82, 82])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ]
        )
    )
    doc.build([Paragraph("Aprilie 2026", styles["Title"]), Spacer(1, 8), table])


def _unpack_attendance_row(row):
    if len(row) == 3:
        name, co_days, mentiuni = row
        return name, co_days, None, None, mentiuni
    if len(row) == 4:
        name, co_days, n_days, mentiuni = row
        return name, co_days, n_days, None, mentiuni
    name, co_days, n_days, cm_days, mentiuni = row
    return name, co_days, n_days, cm_days, mentiuni


def build_attendance_workbook(
    path: Path,
    rows: list[tuple] | None = None,
    title: str = "TRANŞARE - APRILIE 2026",
    period_label: str = "01-30.04.2026",
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "aprilie"
    ws["A1"] = title
    ws.append([])
    ws.append(["Nr. crt.", "Nume și prenume", "CO", "N", "CM", period_label, "Mentiuni"])

    rows = rows or [
        ("CIOCLEA IOAN", "02,06", "07.04 - a plecat la 15:00"),
        ("COJOCARIU SERGIU", None, None),
    ]
    for index, row in enumerate(rows, start=1):
        name, co_days, n_days, cm_days, mentiuni = _unpack_attendance_row(row)
        ws.append([index, name, co_days, n_days, cm_days, "24//26", mentiuni])

    wb.save(path)
    wb.close()


def build_attendance_pdf(
    path: Path,
    rows: list[tuple] | None = None,
    title: str = "TRANSARE - APRILIE 2026",
    period_label: str = "01-30.04.2026",
) -> None:
    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24)
    styles = getSampleStyleSheet()
    table_data = [["Nr. crt.", "Nume si prenume", "CO", "N", "CM", period_label, "Mentiuni"]]
    rows = rows or [
        ("CIOCLEA IOAN", "02,06", None, "07.04 - a plecat la 15:00"),
        ("COJOCARIU SERGIU", None, None, None),
    ]
    for index, row in enumerate(rows, start=1):
        name, co_days, n_days, cm_days, mentiuni = _unpack_attendance_row(row)
        table_data.append([index, name, co_days or "", n_days or "", cm_days or "", "24//26", mentiuni or ""])

    table = Table(table_data, colWidths=[48, 210, 74, 74, 54, 110, 180])
    table.setStyle(
        TableStyle(
            [
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("ALIGN", (1, 1), (1, -1), "LEFT"),
                ("ALIGN", (6, 1), (6, -1), "LEFT"),
            ]
        )
    )
    doc.build([Paragraph(title, styles["Title"]), Spacer(1, 8), table])


def build_target_workbook(
    path: Path,
    include_existing_aprilie: bool = False,
    include_offless_block: bool = False,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "martie"
    populate_template_sheet(ws, "martie", 2026, 3, include_offless_block=include_offless_block)

    ws_other = wb.create_sheet("Sheet1")
    ws_other["A1"] = "aux"

    if include_existing_aprilie:
        stale = wb.create_sheet("aprilie")
        stale["A1"] = "stale"

    wb.save(path)
    wb.close()


def build_basic_target_workbook(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "target"
    wb.save(path)
    wb.close()


def build_predefined_template_workbook(path: Path, include_offless_block: bool = False) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "martie"
    populate_template_sheet(ws, "martie", 2026, 3, include_offless_block=include_offless_block)
    wb.save(path)
    wb.close()


def populate_template_sheet(
    ws,
    month_name: str,
    year: int,
    month_number: int,
    include_offless_block: bool = False,
) -> None:
    ws.title = month_name
    ws.sheet_view.zoomScale = 55

    header_fill = PatternFill(fill_type="solid", fgColor="F2E2D1")
    value_fill = PatternFill(fill_type="solid", fgColor="FFF9F2")
    off_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    bold_font = Font(bold=True)

    day_headers = [float(f"{day.day}.{day.month:02d}") for day in build_target_day_columns(year, month_number)]
    summary_start_col = 3 + len(day_headers)

    for row in range(1, 4):
        ws.row_dimensions[row].height = 10
    ws.row_dimensions[5].height = 22

    for col in range(1, summary_start_col + len(SUMMARY_LABELS) + 2):
        ws.column_dimensions[get_column_letter(col)].width = 11

    blocks = [
        {
            "header_row": 5,
            "id": 1,
            "name": "CIOCLEA IOAN",
            "section": "SEF",
            "base_salary": 9500,
            "mask": [index not in {10, 17, 23} for index in range(len(day_headers))],
            "trans_label": "transatori",
        },
        {
            "header_row": 12,
            "id": 2,
            "name": "COJOCARIU SERGIU",
            "section": "PULPA",
            "base_salary": 8700,
            "mask": (
                [True for _ in range(len(day_headers))]
                if include_offless_block
                else [index not in {0, 6, 12, 18, 24} for index in range(len(day_headers))]
            ),
            "trans_label": None,
        },
    ]

    for block in blocks:
        write_block(
            ws=ws,
            header_row=block["header_row"],
            section=block["section"],
            employee_id=block["id"],
            name=block["name"],
            base_salary=block["base_salary"],
            day_headers=day_headers,
            summary_start_col=summary_start_col,
            attendance_mask=block["mask"],
            header_fill=header_fill,
            value_fill=value_fill,
            off_fill=off_fill,
            bold_font=bold_font,
            trans_label=block["trans_label"],
        )


def write_block(
    ws,
    header_row: int,
    section: str,
    employee_id: int,
    name: str,
    base_salary: int,
    day_headers: list[float],
    summary_start_col: int,
    attendance_mask: list[bool],
    header_fill: PatternFill,
    value_fill: PatternFill,
    off_fill: PatternFill,
    bold_font: Font,
    trans_label: str | None,
) -> None:
    value_row = header_row + 1
    trans_row = header_row + 2
    per_om_row = header_row + 3
    norm_row = header_row + 4
    diff_row = header_row + 5

    ws.cell(header_row, 2, section)
    ws.cell(header_row, 2).font = bold_font
    for offset, day_value in enumerate(day_headers):
        col = 3 + offset
        cell = ws.cell(header_row, col, day_value)
        cell.fill = off_fill if not attendance_mask[offset] else header_fill
        cell.font = bold_font

    for offset, label in enumerate(SUMMARY_LABELS):
        col = summary_start_col + offset
        cell = ws.cell(header_row, col, label)
        cell.fill = header_fill
        cell.font = bold_font

    ws.cell(value_row, 1, employee_id)
    ws.cell(value_row, 2, name)
    ws.cell(trans_row, 2, trans_label)
    ws.cell(per_om_row, 2, "vaci pe om")
    ws.cell(norm_row, 2, "vaci 160 in 14")
    ws.cell(diff_row, 2, "difernta ")

    for offset, present in enumerate(attendance_mask):
        col = 3 + offset
        value = 100 + offset if present else 0
        transatori = 15 if present else 0
        ws.cell(value_row, col, value).fill = value_fill if present else off_fill
        ws.cell(trans_row, col, transatori).fill = value_fill if present else off_fill
        ws.cell(per_om_row, col).fill = value_fill if present else off_fill
        ws.cell(norm_row, col).fill = value_fill if present else off_fill
        ws.cell(diff_row, col).fill = value_fill if present else off_fill
        letter = get_column_letter(col)
        ws.cell(per_om_row, col, f"=IF({letter}{trans_row}>0,{letter}{value_row}/{letter}{trans_row},0)")
        ws.cell(norm_row, col, "=IF(1>0,11.5,0)" if present else 0)
        ws.cell(diff_row, col, f"=IF({letter}{trans_row}>0,{letter}{per_om_row}-{letter}{norm_row},0)")

    sum_col = summary_start_col
    zile_col = sum_col + 1
    target_col = sum_col + 2
    de_platit_col = sum_col + 3
    zile_lucratoare_col = sum_col + 4
    zile_in_plus_col = sum_col + 5
    salar_pe_zi_col = sum_col + 6
    bani_in_plus_col = sum_col + 7
    salar_baza_col = sum_col + 8
    total_col = sum_col + 9
    name_col = sum_col + 10

    first_letter = get_column_letter(3)
    last_letter = get_column_letter(2 + len(day_headers))
    ws.cell(value_row, sum_col, f"=SUM({first_letter}{value_row}:{last_letter}{value_row})")
    ws.cell(value_row, zile_col, f'=COUNTIF({first_letter}{trans_row}:{last_letter}{trans_row},">0")')
    ws.cell(value_row, target_col, f"={get_column_letter(zile_col)}{value_row}*166")
    ws.cell(value_row, de_platit_col, f"={get_column_letter(sum_col)}{value_row}-{get_column_letter(target_col)}{value_row}")
    ws.cell(value_row, zile_lucratoare_col, len(day_headers))
    ws.cell(value_row, zile_in_plus_col, f"={get_column_letter(de_platit_col)}{value_row}/175")
    ws.cell(value_row, salar_pe_zi_col, f"={get_column_letter(salar_baza_col)}{value_row}/{get_column_letter(zile_lucratoare_col)}{value_row}")
    ws.cell(value_row, bani_in_plus_col, f"={get_column_letter(salar_pe_zi_col)}{value_row}*{get_column_letter(zile_in_plus_col)}{value_row}")
    ws.cell(value_row, salar_baza_col, base_salary)
    ws.cell(value_row, total_col, f"={get_column_letter(salar_baza_col)}{value_row}+{get_column_letter(bani_in_plus_col)}{value_row}")
    ws.cell(value_row, name_col, name)
