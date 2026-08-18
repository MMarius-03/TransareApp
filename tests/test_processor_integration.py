from __future__ import annotations

from datetime import date
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import coordinate_to_tuple, range_boundaries
import pytest

import processor
from tests.conftest import (
    build_attendance_pdf,
    build_attendance_workbook,
    build_basic_target_workbook,
    build_predefined_template_workbook,
    build_source_pdf,
    build_source_workbook_for_dates,
    build_source_workbook,
    build_target_workbook,
)


def _day_column(worksheet, layout: processor.SheetLayout, day_number: int) -> int:
    for col in range(layout.day_start_col, layout.day_end_col + 1):
        if processor.parse_sheet_day_header(worksheet.cell(5, col).value) == day_number:
            return col
    raise AssertionError(f"Day {day_number} was not generated")


def _target_rate_cell(worksheet, layout: processor.SheetLayout):
    blocks = processor.detect_sheet_blocks(worksheet, layout)
    return worksheet.cell(blocks[0].header_row - 1, layout.summary_start_col + 2)


def _workday_count_cell(worksheet, layout: processor.SheetLayout):
    blocks = processor.detect_sheet_blocks(worksheet, layout)
    return worksheet.cell(blocks[0].header_row - 1, layout.summary_start_col + 4)


def _formula_cells(worksheet):
    return [
        (cell.coordinate, cell.value)
        for row in worksheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]


def test_run_fill_predefined_mode_creates_new_aprilie_sheet_from_asset(tmp_path: Path, monkeypatch) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    preset_path = tmp_path / "preset.xlsx"
    build_source_workbook(source_path)
    build_basic_target_workbook(target_path)
    build_predefined_template_workbook(preset_path)

    monkeypatch.setattr(processor, "PREDEFINED_TEMPLATE_WORKBOOK", preset_path)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREDEFINED,
        target_month_name="aprilie",
    )

    assert result.output_file.exists()
    assert result.output_file == target_path
    assert result.created_sheet_name == "aprilie"
    assert result.test_sheet_name == "aprilie pentru teste"
    assert result.template_source_name == "martie"
    assert result.mapped_days == 5
    assert len(result.warnings) == 1
    assert "Zile colorate galben pentru că lipsesc din sursă" in result.warnings[0]
    assert "13.04.2026" in result.warnings[0]

    wb = load_workbook(result.output_file, data_only=False)
    assert "aprilie" in wb.sheetnames
    assert "aprilie pentru teste" in wb.sheetnames
    ws = wb["aprilie"]
    layout = processor.get_sheet_layout(ws)

    # All 26 April workdays remain in the table; missing ones are highlighted yellow.
    assert layout.day_end_col - layout.day_start_col + 1 == 26
    assert ws.sheet_view.zoomScale == 55
    assert ws["C5"].fill.fgColor.rgb == "00F2E2D1"
    assert ws["G5"].value == 6.04
    summary_letter = get_column_letter(layout.summary_start_col)
    assert ws.cell(5, layout.summary_start_col).value == "VACI TRAN "
    target_rate_cell = _target_rate_cell(ws, layout)
    assert target_rate_cell.value == 166
    assert target_rate_cell.fill.fgColor.rgb == "00DDEBFF"
    assert target_rate_cell.font.bold is True
    assert target_rate_cell.alignment.horizontal == "center"
    assert ws["C6"].value == 120
    assert ws["D6"].value == "=100+12/6"
    assert ws["D7"].value == 14.67
    assert ws["C13"].value == 120
    zile_col_letter = get_column_letter(layout.summary_start_col + 1)
    last_day_letter = get_column_letter(layout.day_end_col)
    assert ws.cell(6, layout.summary_start_col + 1).value == f'=COUNTIF(C7:{last_day_letter}7,">0")'
    target_col_letter = get_column_letter(layout.summary_start_col + 2)
    target_rate_ref_letter = get_column_letter(target_rate_cell.column)
    assert ws.cell(6, layout.summary_start_col + 2).value == f"={zile_col_letter}6*${target_rate_ref_letter}${target_rate_cell.row}"
    workday_count_cell = _workday_count_cell(ws, layout)
    assert ws.cell(6, layout.summary_start_col + 4).value == f"=${get_column_letter(workday_count_cell.column)}${workday_count_cell.row}"
    assert workday_count_cell.value == 5
    assert workday_count_cell.fill.fgColor.rgb == "00C6EFCE"
    assert workday_count_cell.font.bold is True
    assert workday_count_cell.alignment.horizontal == "center"
    assert wb.sheetnames[0] == "aprilie"

    test_ws = wb["aprilie pentru teste"]
    test_layout = processor.get_sheet_layout(test_ws)
    assert _formula_cells(test_ws) == []
    assert test_ws["D6"].value == 102
    assert test_ws["D7"].value == 14.67
    assert test_ws["D8"].value == pytest.approx(102 / 14.67, rel=0, abs=0.0001)
    assert test_ws.cell(6, test_layout.summary_start_col).value == 614
    assert test_ws.cell(6, test_layout.summary_start_col + 1).value == 5
    assert test_ws["D6"].fill.fgColor.rgb == ws["D6"].fill.fgColor.rgb
    wb.close()


def test_run_fill_raises_processing_cancelled_when_cancel_requested(tmp_path: Path, monkeypatch) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    preset_path = tmp_path / "preset.xlsx"
    build_source_workbook(source_path)
    build_basic_target_workbook(target_path)
    build_predefined_template_workbook(preset_path)
    monkeypatch.setattr(processor, "PREDEFINED_TEMPLATE_WORKBOOK", preset_path)

    try:
        processor.run_fill(
            source_path=source_path,
            target_path=target_path,
            template_mode=processor.TEMPLATE_MODE_PREDEFINED,
            target_month_name="aprilie",
            cancel_check=lambda: True,
        )
    except processor.ProcessingCancelled:
        pass
    else:
        raise AssertionError("Expected run_fill to raise ProcessingCancelled")

    workbook = load_workbook(target_path, data_only=False)
    assert "aprilie" not in workbook.sheetnames
    assert "aprilie pentru teste" not in workbook.sheetnames
    workbook.close()


def test_run_fill_reports_friendly_error_when_target_is_locked(tmp_path: Path, monkeypatch) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    preset_path = tmp_path / "preset.xlsx"
    build_source_workbook(source_path)
    build_basic_target_workbook(target_path)
    build_predefined_template_workbook(preset_path)
    monkeypatch.setattr(processor, "PREDEFINED_TEMPLATE_WORKBOOK", preset_path)

    def _raise_permission(self, *_args, **_kwargs):
        raise PermissionError(13, "Permission denied")

    monkeypatch.setattr(processor.Workbook, "save", _raise_permission)

    try:
        processor.run_fill(
            source_path=source_path,
            target_path=target_path,
            template_mode=processor.TEMPLATE_MODE_PREDEFINED,
            target_month_name="aprilie",
        )
    except processor.ProcessorError as exc:
        message = str(exc)
    else:
        raise AssertionError("Expected run_fill to raise ProcessorError")

    assert "deschis în alt program" in message
    assert "salarii.xlsx" in message


def test_run_fill_keeps_missing_days_as_yellow_columns(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    all_april_days = processor.build_target_day_columns(2026, 4)
    source_days = [
        current_day
        for current_day in all_april_days
        if current_day not in {date(2026, 4, 11), date(2026, 4, 13)}
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    header_values = [
        worksheet.cell(5, col).value
        for col in range(layout.day_start_col, layout.day_end_col + 1)
    ]

    # All 26 April workdays remain (11.04 and 13.04 stay but are highlighted yellow)
    assert len(header_values) == 26
    assert 11.04 in header_values
    assert 13.04 in header_values
    # Missing day columns are painted pure yellow on every row of every block
    missing_11_col = layout.day_start_col + header_values.index(11.04)
    for row in range(6, 11):
        assert worksheet.cell(row, missing_11_col).fill.fgColor.rgb == "00FFFF00"
    assert worksheet.cell(5, layout.summary_start_col).value == "VACI TRAN "
    workday_count_cell = _workday_count_cell(worksheet, layout)
    assert worksheet.cell(6, layout.summary_start_col + 4).value == f"=${get_column_letter(workday_count_cell.column)}${workday_count_cell.row}"
    # workday_count counts only days that actually had transatori in the source (24)
    assert workday_count_cell.value == 24
    assert worksheet.cell(6, layout.summary_start_col + 8).value == 9500
    target_rate_cell = _target_rate_cell(worksheet, layout)
    target_rate_ref = f"${get_column_letter(target_rate_cell.column)}${target_rate_cell.row}"
    zile_col_letter = get_column_letter(layout.summary_start_col + 1)
    zile_lucratoare_ref = f"{get_column_letter(layout.summary_start_col + 4)}6"
    bani_in_plus_ref = f"{get_column_letter(layout.summary_start_col + 7)}6"
    salar_baza_ref = f"{get_column_letter(layout.summary_start_col + 8)}6"
    sum_col_letter = get_column_letter(layout.summary_start_col)
    workday_count_letter = get_column_letter(layout.summary_start_col + 4)
    de_platit_letter = get_column_letter(layout.summary_start_col + 3)
    target_col_letter = get_column_letter(layout.summary_start_col + 2)
    last_day_letter = get_column_letter(layout.day_end_col)
    assert target_rate_cell.value == 166
    assert target_rate_cell.fill.fgColor.rgb == "00DDEBFF"
    assert worksheet.cell(6, layout.summary_start_col + 2).value == f"={zile_col_letter}6*{target_rate_ref}"
    assert worksheet.cell(6, layout.summary_start_col + 6).value == f"={salar_baza_ref}/{zile_lucratoare_ref}"
    assert worksheet.cell(6, layout.summary_start_col + 9).value == f"={salar_baza_ref}+{bani_in_plus_ref}"
    assert worksheet.cell(8, layout.summary_start_col).value == f"=SUM(C8:{last_day_letter}8)"
    assert worksheet.cell(9, layout.summary_start_col).value == f"=SUM(C9:{last_day_letter}9)"
    assert worksheet.cell(10, layout.summary_start_col).value == f"=SUM(C10:{last_day_letter}10)"
    assert worksheet.cell(8, layout.summary_start_col + 2).value == "cost /vaca"
    assert worksheet.cell(9, layout.summary_start_col + 2).value == "bani  in plus"
    salar_pe_zi_letter = get_column_letter(layout.summary_start_col + 6)
    assert worksheet.cell(8, layout.summary_start_col + 4).value == f"={salar_pe_zi_letter}6/11.5"
    assert worksheet.cell(9, layout.summary_start_col + 4).value == f"={sum_col_letter}10*{workday_count_letter}8"
    assert worksheet.cell(10, layout.summary_start_col + 5).value == f"={de_platit_letter}10/175"
    assert result.warnings == [
        "Zile colorate galben pentru că lipsesc din sursă: 11.04.2026, 13.04.2026"
    ]

    workbook.close()


def test_run_fill_can_exclude_missing_days_from_the_generated_report(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    all_april_days = processor.build_target_day_columns(2026, 4)
    source_days = [
        current_day
        for current_day in all_april_days
        if current_day not in {date(2026, 4, 11), date(2026, 4, 13)}
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        include_missing_source_days=False,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    header_values = [
        worksheet.cell(5, col).value
        for col in range(layout.day_start_col, layout.day_end_col + 1)
    ]

    assert len(header_values) == 24
    assert 11.04 not in header_values
    assert 13.04 not in header_values
    assert _workday_count_cell(worksheet, layout).value == 24
    assert result.warnings == [
        "Zile neincluse în tabel pentru că lipsesc din sursă: 11.04.2026, 13.04.2026"
    ]

    workbook.close()


def test_run_fill_can_treat_wrong_month_source_date_as_target_month(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    source_days = processor.build_target_day_columns(2026, 4)
    build_source_workbook_for_dates(source_path, source_days)
    workbook = load_workbook(source_path)
    worksheet = workbook.active
    worksheet["A8"] = "07.03.2026"
    workbook.save(source_path)
    workbook.close()
    build_target_workbook(target_path)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        treat_mismatched_source_dates_as_target=True,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    day_seven_col = _day_column(worksheet, layout, 7)

    assert worksheet.cell(6, day_seven_col).value == 106
    assert worksheet.cell(7, day_seven_col).value == 15
    assert any("07.03.2026 → 07.04.2026" in warning for warning in result.warnings)

    workbook.close()


def test_run_fill_previous_sheet_mode_overwrites_stale_global_target_rate(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)

    workbook = load_workbook(target_path, data_only=False)
    worksheet = workbook["martie"]
    layout = processor.get_sheet_layout(worksheet)
    blocks = processor.detect_sheet_blocks(worksheet, layout)
    stale_cell = worksheet.cell(blocks[0].header_row - 1, layout.summary_start_col + 2)
    stale_cell.value = 123
    workbook.save(target_path)
    workbook.close()

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    target_rate_cell = _target_rate_cell(worksheet, layout)
    target_formula = worksheet.cell(6, layout.summary_start_col + 2).value

    assert target_rate_cell.value == 166
    assert target_rate_cell.fill.fgColor.rgb == "00DDEBFF"
    zile_col_letter = get_column_letter(layout.summary_start_col + 1)
    assert target_formula == f"={zile_col_letter}6*${get_column_letter(target_rate_cell.column)}${target_rate_cell.row}"
    assert "*166" not in target_formula

    workbook.close()


def test_run_fill_does_not_generate_formula_ranges_that_include_their_own_cell(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    all_april_days = processor.build_target_day_columns(2026, 4)
    source_days = [
        current_day
        for current_day in all_april_days
        if current_day not in {date(2026, 4, 11), date(2026, 4, 13)}
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]

    bad_formulas = []
    ref_errors = []
    for row in worksheet.iter_rows():
        for cell in row:
            formula = cell.value
            if not (isinstance(formula, str) and formula.startswith("=")):
                continue
            if "#REF!" in formula:
                ref_errors.append((cell.coordinate, formula))
            cell_row, cell_col = coordinate_to_tuple(cell.coordinate)
            for match in re.finditer(r"([A-Z]+\d+):([A-Z]+\d+)", formula):
                min_col, min_row, max_col, max_row = range_boundaries(match.group(0))
                if min_row <= cell_row <= max_row and min_col <= cell_col <= max_col:
                    bad_formulas.append((cell.coordinate, formula))

    assert ref_errors == []
    assert bad_formulas == []

    workbook.close()


def test_run_fill_ignores_attendance_co_days_for_days_missing_from_source(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    all_april_days = processor.build_target_day_columns(2026, 4)
    source_days = [
        current_day
        for current_day in all_april_days
        if current_day not in {date(2026, 4, 11), date(2026, 4, 13)}
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)
    build_attendance_workbook(
        attendance_path,
        rows=[("CIOCLEA IOAN", "11,13", "11.04 - zi eliminată")],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    header_values = [
        worksheet.cell(5, col).value
        for col in range(layout.day_start_col, layout.day_end_col + 1)
    ]

    # Missing-from-source days remain in the header (colored yellow), but the CO from
    # attendance is skipped for them because the cell is already the zero/zero pair.
    assert 11.04 in header_values
    assert 13.04 in header_values
    assert result.attendance_summary is not None
    assert result.attendance_summary.co_days_applied == 0
    assert result.attendance_summary.mentiuni_days_colored == 0

    workbook.close()


def test_run_fill_predefined_mode_colors_zero_result_block_using_global_off_fallback(
    tmp_path: Path,
    monkeypatch,
) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    preset_path = tmp_path / "preset.xlsx"
    build_source_workbook(source_path, zero_days={"02.04.2026"})
    build_basic_target_workbook(target_path)
    build_predefined_template_workbook(preset_path, include_offless_block=True)

    monkeypatch.setattr(processor, "PREDEFINED_TEMPLATE_WORKBOOK", preset_path)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREDEFINED,
        target_month_name="aprilie",
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]

    for row in (12, 13, 14, 15, 16, 17):
        assert worksheet.cell(row, 4).fill.fgColor.rgb == "00FFF2CC"
    assert worksheet["D13"].value == 0
    assert worksheet["D14"].value == 0

    workbook.close()


def test_run_fill_previous_sheet_mode_numbers_existing_target_and_inserts_before_template(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path, include_existing_aprilie=True)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    wb = load_workbook(result.output_file, data_only=False)
    assert result.created_sheet_name == "aprilie (2)"
    assert result.test_sheet_name == "aprilie pentru teste (2)"
    assert wb.sheetnames[:3] == ["aprilie (2)", "aprilie pentru teste (2)", "martie"]
    assert wb["aprilie"]["A1"].value == "stale"

    ws = wb["aprilie (2)"]
    assert ws["A1"].value is None
    assert ws["G12"].value == 6.04
    layout_ws = processor.get_sheet_layout(ws)
    assert ws.cell(5, layout_ws.summary_start_col).value == "VACI TRAN "
    salar_baza_letter = get_column_letter(layout_ws.summary_start_col + 8)
    bani_in_plus_letter = get_column_letter(layout_ws.summary_start_col + 7)
    total_col = layout_ws.summary_start_col + 9
    assert ws.cell(6, total_col).value == f"={salar_baza_letter}6+{bani_in_plus_letter}6"
    assert _formula_cells(wb["aprilie pentru teste (2)"]) == []
    wb.close()


def test_run_fill_numbers_pair_when_test_sheet_name_exists(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)

    workbook = load_workbook(target_path, data_only=False)
    workbook.create_sheet("aprilie pentru teste")
    workbook.save(target_path)
    workbook.close()

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    workbook = load_workbook(result.output_file, data_only=False)
    assert result.created_sheet_name == "aprilie (2)"
    assert result.test_sheet_name == "aprilie pentru teste (2)"
    assert "aprilie" not in workbook.sheetnames
    assert "aprilie pentru teste" in workbook.sheetnames
    assert "aprilie (2)" in workbook.sheetnames
    assert "aprilie pentru teste (2)" in workbook.sheetnames
    workbook.close()


def test_run_fill_creates_month_and_pentru_teste_sheet_names(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    source_days = processor.build_target_day_columns(2026, 5)[:3]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="mai",
        template_sheet_name="martie",
    )

    workbook = load_workbook(result.output_file, data_only=False)
    assert result.created_sheet_name == "mai"
    assert result.test_sheet_name == "mai pentru teste"
    assert "mai" in workbook.sheetnames
    assert "mai pentru teste" in workbook.sheetnames
    assert _formula_cells(workbook["mai pentru teste"]) == []
    workbook.close()


def test_run_fill_previous_sheet_mode_keeps_normal_fill_for_zero_result_blocks(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    build_source_workbook(source_path, zero_days={"02.04.2026"})
    build_target_workbook(target_path, include_offless_block=True)

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]

    assert worksheet["D13"].value == 0
    assert worksheet["D14"].value == 0
    assert worksheet["D12"].fill.fgColor.rgb == "00F2E2D1"
    for cell_ref in ("D13", "D14", "D15", "D16", "D17"):
        assert worksheet[cell_ref].fill.fgColor.rgb == "00FFF9F2"

    workbook.close()


def test_run_fill_with_attendance_applies_co_days_and_mentiuni(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)
    build_attendance_workbook(
        attendance_path,
        rows=[("CIOCLEA IOAN", "02,06", "07.04 - a plecat la 15:00")],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout_apr = processor.get_sheet_layout(worksheet)
    mentiuni_col_letter = get_column_letter(layout_apr.summary_start_col + 12)

    assert worksheet["D6"].value == 0
    assert worksheet["D7"].value == 0
    assert worksheet["G6"].value == 0
    assert worksheet["G7"].value == 0
    assert worksheet["D8"].value == "=IF(D7>0,D6/D7,0)"
    assert worksheet["D9"].value == "=IF(D7>0,11.5,0)"
    assert worksheet["D10"].value == "=IF(D7>0,D8-D9,0)"
    assert worksheet["D6"].fill.fgColor.rgb == "00FFFF00"
    assert worksheet[f"{mentiuni_col_letter}5"].value == "Mentiuni"
    assert worksheet[f"{mentiuni_col_letter}6"].value == "07.04 - a plecat la 15:00"
    # Mentiuni note cell inherits the "off" fill of the row, which is now pure yellow
    # (since missing-source days paint the block yellow).
    assert worksheet[f"{mentiuni_col_letter}6"].fill.fgColor.rgb == "00FFFF00"
    assert worksheet[f"{mentiuni_col_letter}13"].value is None
    assert worksheet.row_dimensions[6].height >= 18
    assert result.attendance_summary is not None
    assert result.attendance_summary.matched_employees == 1
    assert result.attendance_summary.co_days_applied == 2
    # Day 7 now stays in the table (yellow, missing from source) so the mentiuni
    # day-fill is applied instead of skipped.
    assert result.attendance_summary.mentiuni_days_colored == 1
    assert result.attendance_summary.mentiuni_copied == 1

    workbook.close()


def test_run_fill_with_attendance_applies_n_days_as_yellow_zero_absences(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    source_days = [
        date(2026, 4, 1),
        date(2026, 4, 2),
        date(2026, 4, 3),
        date(2026, 4, 4),
        date(2026, 4, 6),
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)
    build_attendance_workbook(
        attendance_path,
        rows=[("CIOCLEA IOAN", "02", "03", "04.04 - observatie")],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    co_col = _day_column(worksheet, layout, 2)
    n_col = _day_column(worksheet, layout, 3)
    mentiuni_col = _day_column(worksheet, layout, 4)

    assert worksheet.cell(6, co_col).value == 0
    assert worksheet.cell(7, co_col).value == 0
    assert worksheet.cell(6, n_col).value == 0
    assert worksheet.cell(7, n_col).value == 0
    for row in (5, 6, 7, 8, 9, 10):
        assert worksheet.cell(row, n_col).fill.fgColor.rgb == "00FFFF00"
    for row in (5, 6, 7, 8, 9, 10):
        assert worksheet.cell(row, mentiuni_col).fill.fgColor.rgb == "00FFF8D6"
    assert result.attendance_summary is not None
    assert result.attendance_summary.co_days_applied == 1
    assert result.attendance_summary.n_days_applied == 1
    assert result.attendance_summary.mentiuni_days_colored == 1

    workbook.close()


def test_run_fill_with_attendance_applies_cm_days_as_yellow_zero_absences(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    source_days = [
        date(2026, 4, 1),
        date(2026, 4, 2),
        date(2026, 4, 3),
        date(2026, 4, 4),
        date(2026, 4, 6),
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)
    build_attendance_workbook(
        attendance_path,
        rows=[("CIOCLEA IOAN", None, None, "04", None)],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    cm_col = _day_column(worksheet, layout, 4)

    assert worksheet.cell(6, cm_col).value == 0
    assert worksheet.cell(7, cm_col).value == 0
    for row in (5, 6, 7, 8, 9, 10):
        assert worksheet.cell(row, cm_col).fill.fgColor.rgb == "00FFFF00"
    assert result.attendance_summary is not None
    assert result.attendance_summary.cm_days_applied == 1

    workbook.close()


def test_run_fill_accepts_pdf_source_and_pdf_attendance(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.pdf"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.pdf"
    build_source_pdf(source_path)
    build_target_workbook(target_path)
    build_attendance_pdf(
        attendance_path,
        rows=[("CIOCLEA IOAN", "02", "03", "04.04 - observatie")],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    n_col = _day_column(worksheet, layout, 3)

    assert worksheet.cell(6, n_col).value == 0
    assert worksheet.cell(7, n_col).value == 0
    assert result.mapped_days == 5
    assert result.attendance_summary is not None
    assert result.attendance_summary.n_days_applied == 1

    workbook.close()


def test_run_fill_with_attendance_colors_mentiuni_day_pale_yellow_without_changing_values(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    source_days = [
        date(2026, 4, 1),
        date(2026, 4, 2),
        date(2026, 4, 3),
        date(2026, 4, 4),
        date(2026, 4, 6),
        date(2026, 4, 7),
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)
    build_attendance_workbook(
        attendance_path,
        rows=[("CIOCLEA IOAN", None, "07.04 - a plecat la 15:00")],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    day_col = _day_column(worksheet, layout, 7)
    day_letter = get_column_letter(day_col)

    assert worksheet.cell(6, day_col).value != 0
    assert worksheet.cell(7, day_col).value != 0
    assert worksheet.cell(8, day_col).value == f"=IF({day_letter}7>0,{day_letter}6/{day_letter}7,0)"
    assert worksheet.cell(9, day_col).value == f"=IF({day_letter}7>0,11.5,0)"
    assert worksheet.cell(10, day_col).value == f"=IF({day_letter}7>0,{day_letter}8-{day_letter}9,0)"
    for row in (5, 6, 7, 8, 9, 10):
        assert worksheet.cell(row, day_col).fill.fgColor.rgb == "00FFF8D6"

    mentiuni_col = layout.summary_start_col + len(processor.SUMMARY_LABELS)
    assert worksheet.cell(5, mentiuni_col).value == "Mentiuni"
    assert worksheet.cell(6, mentiuni_col).value == "07.04 - a plecat la 15:00"
    assert result.attendance_summary is not None
    assert result.attendance_summary.mentiuni_days_colored == 1
    assert result.attendance_summary.mentiuni_copied == 1

    workbook.close()


def test_run_fill_with_attendance_keeps_co_fill_when_co_and_mentiuni_overlap(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    source_days = [
        date(2026, 4, 1),
        date(2026, 4, 2),
        date(2026, 4, 3),
        date(2026, 4, 4),
        date(2026, 4, 6),
        date(2026, 4, 7),
    ]
    build_source_workbook_for_dates(source_path, source_days)
    build_target_workbook(target_path)
    build_attendance_workbook(
        attendance_path,
        rows=[("CIOCLEA IOAN", "07", "07.04 - CO și mențiune")],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout = processor.get_sheet_layout(worksheet)
    day_col = _day_column(worksheet, layout, 7)

    assert worksheet.cell(6, day_col).value == 0
    assert worksheet.cell(7, day_col).value == 0
    for row in (5, 6, 7, 8, 9, 10):
        assert worksheet.cell(row, day_col).fill.fgColor.rgb == "00FFFF00"
    assert result.attendance_summary is not None
    assert result.attendance_summary.co_days_applied == 1
    assert result.attendance_summary.mentiuni_days_colored == 0

    workbook.close()


def test_run_fill_with_attendance_inserts_mentiuni_when_neighbor_column_is_not_blank(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)
    build_attendance_workbook(attendance_path)

    workbook = load_workbook(target_path, data_only=False)
    worksheet = workbook["martie"]
    layout = processor.get_sheet_layout(worksheet)
    premium_col = layout.summary_start_col + len(processor.SUMMARY_LABELS) - 1
    worksheet.cell(5, premium_col + 1).value = "RESERVED"
    workbook.save(target_path)
    workbook.close()

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout_apr = processor.get_sheet_layout(worksheet)
    mentiuni_col_letter = get_column_letter(layout_apr.summary_start_col + 12)
    reserved_col_letter = get_column_letter(layout_apr.summary_start_col + 13)

    assert worksheet[f"{mentiuni_col_letter}5"].value == "Mentiuni"
    assert worksheet[f"{reserved_col_letter}5"].value == "RESERVED"

    workbook.close()


def test_run_fill_with_attendance_uses_approximate_matching_and_logs_unmatched(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)

    workbook = load_workbook(target_path, data_only=False)
    worksheet = workbook["martie"]
    worksheet["B13"] = "NECHIFOR NELU FLORIN"
    workbook.save(target_path)
    workbook.close()

    build_attendance_workbook(
        attendance_path,
        rows=[
            ("NICHIFOR NELU-FLORIN", "02", "notă aproximativă"),
            ("PERSOANA LIPSA", "02", None),
        ],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    workbook = load_workbook(result.output_file, data_only=False)
    worksheet = workbook["aprilie"]
    layout_apr = processor.get_sheet_layout(worksheet)
    blocks_apr = processor.detect_sheet_blocks(worksheet, layout_apr)
    mentiuni_col_letter = get_column_letter(layout_apr.summary_start_col + 12)
    # Find the NECHIFOR block (approximate match target) by name
    nechifor_block = next(
        b for b in blocks_apr
        if processor.normalize_person_name(worksheet.cell(b.value_row, 2).value or "")
        == processor.normalize_person_name("NECHIFOR NELU FLORIN")
    )
    assert worksheet.cell(nechifor_block.value_row, 4).value == 0  # day 2.04 -> CO
    assert worksheet.cell(nechifor_block.value_row, layout_apr.summary_start_col + 12).value == "notă aproximativă"
    assert result.attendance_summary is not None
    assert result.attendance_summary.approximate_matches == 1
    # PERSOANA LIPSA is now automatically added as a new employee row instead of being
    # left in unmatched_names, so unmatched is empty and added_employees contains it.
    assert result.attendance_summary.unmatched_names == []
    assert result.attendance_summary.added_employees == ["PERSOANA LIPSA"]
    named_blocks = [block for block in blocks_apr if worksheet.cell(block.value_row, 2).value]
    assert [worksheet.cell(block.value_row, 1).value for block in named_blocks] == list(
        range(1, len(named_blocks) + 1)
    )

    workbook.close()


def test_run_fill_with_attendance_logs_ambiguous_names(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)

    workbook = load_workbook(target_path, data_only=False)
    worksheet = workbook["martie"]
    worksheet["B6"] = "POPESCU ION VASILE"
    worksheet["B13"] = "IONESCU ION VASILE"
    workbook.save(target_path)
    workbook.close()

    build_attendance_workbook(
        attendance_path,
        rows=[("ION VASILE", "02", "notă ambiguă")],
    )

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
    )

    assert result.attendance_summary is not None
    assert len(result.attendance_summary.ambiguous_names) == 1
    assert "POPESCU ION VASILE" in result.attendance_summary.ambiguous_names[0]
    assert "IONESCU ION VASILE" in result.attendance_summary.ambiguous_names[0]


def test_run_fill_with_attendance_month_mismatch_blocks_by_default_and_can_continue(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)
    build_attendance_workbook(
        attendance_path,
        title="TRANŞARE - MARTIE 2026",
        period_label="01-31.03.2026",
    )

    try:
        processor.run_fill(
            source_path=source_path,
            target_path=target_path,
            template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
            target_month_name="aprilie",
            template_sheet_name="martie",
            attendance_path=attendance_path,
        )
    except processor.ProcessorError as exc:
        assert "Luna din Transatori+Detinuți" in str(exc)
    else:
        raise AssertionError("Expected ProcessorError")

    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
        attendance_path=attendance_path,
        allow_attendance_month_mismatch=True,
    )

    assert result.output_file.exists()


def test_run_fill_sorts_employee_blocks_alphabetically(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)

    workbook = load_workbook(target_path)
    worksheet = workbook["martie"]
    worksheet["B6"] = "ZETU ION"
    worksheet["B13"] = "ALBU ION"
    workbook.save(target_path)
    workbook.close()

    processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    workbook = load_workbook(target_path, data_only=False)
    worksheet = workbook["aprilie"]
    blocks = processor.detect_sheet_blocks(worksheet, None)
    names = [worksheet.cell(block.value_row, 2).value for block in blocks]
    workbook.close()

    assert names[:2] == ["ALBU ION", "ZETU ION"]


def test_run_fill_validates_source_month_mismatch(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path)

    try:
        processor.run_fill(
            source_path=source_path,
            target_path=target_path,
            template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
            target_month_name="martie",
            template_sheet_name="martie",
        )
    except processor.ProcessorError as exc:
        message = str(exc)
        assert "aprilie" in message and "martie" in message
    else:
        raise AssertionError("Expected ProcessorError")


def test_run_fill_proceeds_with_selected_month_when_detection_fails(tmp_path: Path, monkeypatch) -> None:
    source_path = tmp_path / "situatie.xlsx"
    target_path = tmp_path / "salarii.xlsx"
    build_source_workbook(source_path)
    build_target_workbook(target_path, include_existing_aprilie=True)

    def _detection_fails(*args, **kwargs):
        raise processor.ProcessorError("Nu am putut detecta luna din fișierul sursă.")

    monkeypatch.setattr(processor, "detect_source_month_from_file", _detection_fails)

    # Even when automatic month detection is broken, the user's explicit month choice
    # must drive the run instead of aborting.
    result = processor.run_fill(
        source_path=source_path,
        target_path=target_path,
        template_mode=processor.TEMPLATE_MODE_PREVIOUS_SHEET,
        target_month_name="aprilie",
        template_sheet_name="martie",
    )

    assert result.output_file.exists()
    assert result.created_sheet_name == "aprilie (2)"
    assert result.test_sheet_name == "aprilie pentru teste (2)"
