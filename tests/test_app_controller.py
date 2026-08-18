from __future__ import annotations

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pytest
from openpyxl import load_workbook

import app_controller
import processor
from app_window import TEMPLATE_MODE_PREDEFINED_LABEL, TEMPLATE_MODE_PREVIOUS_SHEET_LABEL
from tests.conftest import (
    build_attendance_workbook,
    build_source_workbook,
    build_source_workbook_for_dates,
    build_target_workbook,
)


class DummySignal:
    def __init__(self) -> None:
        self.callbacks = []

    def connect(self, callback) -> None:
        self.callbacks.append(callback)


class FakeWindow:
    def __init__(self) -> None:
        self.browse_source_requested = DummySignal()
        self.browse_target_requested = DummySignal()
        self.browse_attendance_requested = DummySignal()
        self.run_requested = DummySignal()
        self.cancel_requested = DummySignal()
        self.open_output_requested = DummySignal()
        self.values = {
            "source_path": "",
            "target_path": "",
            "attendance_path": "",
            "target_month_name": "ianuarie",
            "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
            "template_sheet_name": "",
        }
        self.status_calls: list[tuple[str, str]] = []
        self.log_messages: list[str] = []
        self.busy_calls: list[tuple[bool, bool]] = []
        self.progress_calls: list[tuple[float, str]] = []
        self.open_enabled = False

    def form_values(self) -> dict[str, str]:
        return dict(self.values)

    def set_source_path(self, path: str) -> None:
        self.values["source_path"] = path

    def set_target_path(self, path: str) -> None:
        self.values["target_path"] = path

    def set_attendance_path(self, path: str) -> None:
        self.values["attendance_path"] = path

    def set_target_month(self, month_name: str) -> None:
        self.values["target_month_name"] = month_name

    def set_template_sheet_options(self, names: list[str], selected: str = "") -> None:
        self.values["template_sheet_name"] = selected or (names[0] if names else "")

    def set_sheet_options(self, names: list[str], selected: str = "") -> None:
        self.set_template_sheet_options(names, selected)

    def set_status(self, message: str, kind: str = "info") -> None:
        self.status_calls.append((message, kind))

    def append_log(self, message: str) -> None:
        self.log_messages.append(message)

    def set_busy(self, is_busy: bool, last_output_exists: bool = False) -> None:
        self.busy_calls.append((is_busy, last_output_exists))

    def set_progress(self, percent: float, text: str = "") -> None:
        self.progress_calls.append((percent, text))

    def set_open_output_enabled(self, enabled: bool) -> None:
        self.open_enabled = enabled


@dataclass
class FakeResult:
    output_file: Path
    created_sheet_name: str = "aprilie"
    test_sheet_name: str = "aprilie pentru teste"
    template_source_name: str | None = "martie"
    mapped_days: int = 5
    updated_blocks: int = 2
    warnings: list[str] | None = None
    differences: list[str] | None = None


class FakeWorker:
    last_kwargs: dict | None = None

    def __init__(self, **kwargs) -> None:
        FakeWorker.last_kwargs = kwargs
        self.progress = DummySignal()
        self.success = DummySignal()
        self.error = DummySignal()
        self.cancelled = DummySignal()
        self.finished = DummySignal()
        self.cancel_count = 0

    def cancel(self) -> None:
        self.cancel_count += 1

    def moveToThread(self, _thread) -> None:
        return None

    def deleteLater(self) -> None:
        return None

    def run(self) -> None:
        return None


class FakeThread:
    def __init__(self, *_args, **_kwargs) -> None:
        self.started = DummySignal()
        self.started_count = 0

    def start(self) -> None:
        self.started_count += 1

    def quit(self) -> None:
        return None

    def wait(self, _timeout: int) -> None:
        return None

    def deleteLater(self) -> None:
        return None


@pytest.fixture
def controller(tmp_path: Path) -> app_controller.AppController:
    window = FakeWindow()
    ctrl = app_controller.AppController(window, base_dir=tmp_path)
    ctrl._show_warning = lambda message: None
    ctrl._show_error = lambda message: None
    return ctrl


def test_initialize_sets_ready_state(controller: app_controller.AppController) -> None:
    controller.initialize()
    assert controller.window.busy_calls[-1] == (False, False)


def test_browse_source_file_detects_month(controller: app_controller.AppController, tmp_path: Path, monkeypatch) -> None:
    source = tmp_path / "source.xlsx"
    build_source_workbook(source)
    monkeypatch.setattr(app_controller.QFileDialog, "getOpenFileName", lambda *_args, **_kwargs: (str(source), ""))

    controller.browse_source_file()

    assert controller.window.values["source_path"] == str(source)
    assert controller.window.values["target_month_name"] == "aprilie"
    assert controller.detected_source_month_name == "aprilie"


def test_browse_target_file_prefers_previous_month_sheet(controller: app_controller.AppController, tmp_path: Path, monkeypatch) -> None:
    target = tmp_path / "target.xlsx"
    build_target_workbook(target)
    controller.window.values["target_month_name"] = "aprilie"
    monkeypatch.setattr(app_controller.QFileDialog, "getOpenFileName", lambda *_args, **_kwargs: (str(target), ""))

    controller.browse_target_file()

    assert controller.window.values["target_path"] == str(target)
    assert controller.window.values["template_sheet_name"] == "martie"


def test_browse_attendance_file_sets_optional_path(controller: app_controller.AppController, tmp_path: Path, monkeypatch) -> None:
    attendance = tmp_path / "attendance.xlsx"
    build_attendance_workbook(attendance)
    monkeypatch.setattr(app_controller.QFileDialog, "getOpenFileName", lambda *_args, **_kwargs: (str(attendance), ""))

    controller.browse_attendance_file()

    assert controller.window.values["attendance_path"] == str(attendance)


def test_start_run_validates_missing_inputs(controller: app_controller.AppController) -> None:
    warnings: list[str] = []
    controller._show_warning = lambda message: warnings.append(message)

    controller.start_run()

    assert warnings == ["Completează toate câmpurile înainte de rulare."]


def test_start_run_requires_template_sheet_in_previous_mode(controller: app_controller.AppController, tmp_path: Path) -> None:
    warnings: list[str] = []
    controller._show_warning = lambda message: warnings.append(message)
    controller.window.values = {
        "source_path": str(tmp_path / "source.xlsx"),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREVIOUS_SHEET_LABEL,
        "template_sheet_name": "",
    }

    controller.start_run()

    assert warnings == ["Alege un sheet template pentru modul de copiere din sheet anterior."]


def test_start_run_creates_worker_with_new_contract(controller: app_controller.AppController, tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    controller.window.values = {
        "source_path": str(tmp_path / "source.xlsx"),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert controller.window.busy_calls[-1][0] is True
    assert isinstance(controller._thread, FakeThread)
    assert FakeWorker.last_kwargs == {
        "source_path": str(tmp_path / "source.xlsx"),
        "target_path": str(tmp_path / "target.xlsx"),
        "template_mode": app_controller.TEMPLATE_MODE_PREDEFINED,
        "target_month_name": "aprilie",
        "template_sheet_name": "martie",
        "attendance_path": None,
        "allow_attendance_month_mismatch": False,
        "include_missing_source_days": True,
        "treat_mismatched_source_dates_as_target": False,
    }


def test_start_run_can_continue_after_attendance_month_mismatch(
    controller: app_controller.AppController,
    tmp_path: Path,
    monkeypatch,
) -> None:
    attendance = tmp_path / "attendance.xlsx"
    build_attendance_workbook(
        attendance,
        title="TRANŞARE - MARTIE 2026",
        period_label="01-31.03.2026",
    )
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    controller._confirm_attendance_month_mismatch = lambda *_args: True
    controller.window.values = {
        "source_path": str(tmp_path / "source.xlsx"),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": str(attendance),
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert FakeWorker.last_kwargs is not None
    assert FakeWorker.last_kwargs["attendance_path"] == str(attendance)
    assert FakeWorker.last_kwargs["allow_attendance_month_mismatch"] is True


def test_start_run_warns_and_can_continue_when_source_days_are_missing(
    controller: app_controller.AppController,
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "source.xlsx"
    build_source_workbook(source)
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    confirmations: list[list[date]] = []
    controller._choose_missing_source_days = lambda missing_days: confirmations.append(missing_days) or True
    controller.window.values = {
        "source_path": str(source),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert confirmations == [[date(2026, 4, 7), date(2026, 4, 8), date(2026, 4, 9), date(2026, 4, 10), date(2026, 4, 11), date(2026, 4, 13), date(2026, 4, 14), date(2026, 4, 15), date(2026, 4, 16), date(2026, 4, 17), date(2026, 4, 18), date(2026, 4, 20), date(2026, 4, 21), date(2026, 4, 22), date(2026, 4, 23), date(2026, 4, 24), date(2026, 4, 25), date(2026, 4, 27), date(2026, 4, 28), date(2026, 4, 29), date(2026, 4, 30)]]
    assert isinstance(controller._thread, FakeThread)
    assert FakeWorker.last_kwargs is not None
    assert FakeWorker.last_kwargs["include_missing_source_days"] is True
    assert any("Zile adăugate cu 0" in message for message in controller.window.log_messages)


def test_start_run_stops_when_user_declines_missing_source_days(
    controller: app_controller.AppController,
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    build_source_workbook(source)
    controller._choose_missing_source_days = lambda _missing_days: None
    controller.window.values = {
        "source_path": str(source),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert controller._thread is None
    assert controller._worker is None
    assert controller.window.log_messages[-1] == "Procesarea a fost oprită deoarece lipsesc zile din fișierul sursă."


def test_start_run_can_exclude_missing_source_days_from_report(
    controller: app_controller.AppController,
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "source.xlsx"
    build_source_workbook(source)
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    controller._choose_missing_source_days = lambda _missing_days: False
    controller.window.values = {
        "source_path": str(source),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert FakeWorker.last_kwargs is not None
    assert FakeWorker.last_kwargs["include_missing_source_days"] is False
    assert any("Zile neincluse în raport" in message for message in controller.window.log_messages)


def test_start_run_can_treat_wrong_month_source_dates_as_selected_month(
    controller: app_controller.AppController,
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "source.xlsx"
    source_dates = processor.build_target_day_columns(2026, 4)
    build_source_workbook_for_dates(source, source_dates)
    workbook = load_workbook(source)
    workbook.active["A8"] = "07.03.2026"
    workbook.save(source)
    workbook.close()
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    confirmations: list[list[tuple[date, date]]] = []
    controller._choose_mismatched_source_dates = lambda mismatches: confirmations.append(mismatches) or True
    controller.window.values = {
        "source_path": str(source),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert confirmations == [[(date(2026, 3, 7), date(2026, 4, 7))]]
    assert FakeWorker.last_kwargs is not None
    assert FakeWorker.last_kwargs["treat_mismatched_source_dates_as_target"] is True
    assert any("07.03.2026 → 07.04.2026" in message for message in controller.window.log_messages)


def test_start_run_stops_when_user_declines_wrong_month_source_dates(
    controller: app_controller.AppController,
    tmp_path: Path,
) -> None:
    source = tmp_path / "source.xlsx"
    source_dates = processor.build_target_day_columns(2026, 4)
    build_source_workbook_for_dates(source, source_dates)
    workbook = load_workbook(source)
    workbook.active["A8"] = "07.03.2026"
    workbook.save(source)
    workbook.close()
    controller._choose_mismatched_source_dates = lambda _mismatches: None
    controller.window.values = {
        "source_path": str(source),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert controller._thread is None
    assert controller.window.log_messages[-1] == (
        "Procesarea a fost oprită deoarece există date din altă lună în fișierul sursă."
    )


def test_start_run_can_keep_wrong_month_source_dates_unchanged(
    controller: app_controller.AppController,
    tmp_path: Path,
    monkeypatch,
) -> None:
    source = tmp_path / "source.xlsx"
    source_dates = processor.build_target_day_columns(2026, 4)
    build_source_workbook_for_dates(source, source_dates)
    workbook = load_workbook(source)
    workbook.active["A8"] = "07.03.2026"
    workbook.save(source)
    workbook.close()
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    controller._choose_mismatched_source_dates = lambda _mismatches: False
    controller._choose_missing_source_days = lambda _missing_days: False
    controller.window.values = {
        "source_path": str(source),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert FakeWorker.last_kwargs is not None
    assert FakeWorker.last_kwargs["treat_mismatched_source_dates_as_target"] is False
    assert FakeWorker.last_kwargs["include_missing_source_days"] is False


def test_start_run_prompts_on_attendance_year_mismatch_with_same_month(
    controller: app_controller.AppController,
    tmp_path: Path,
    monkeypatch,
) -> None:
    attendance = tmp_path / "attendance.xlsx"
    # Same month as target (aprilie) but a different year than the detected source.
    build_attendance_workbook(
        attendance,
        title="TRANŞARE - APRILIE 2025",
        period_label="01-30.04.2025",
    )
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    controller.detected_source_year = 2026
    confirm_calls: list[tuple[str, str]] = []

    def _confirm(attendance_month: str, target_month_name: str) -> bool:
        confirm_calls.append((attendance_month, target_month_name))
        return True

    controller._confirm_attendance_month_mismatch = _confirm
    controller.window.values = {
        "source_path": str(tmp_path / "source.xlsx"),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": str(attendance),
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }

    controller.start_run()

    assert confirm_calls, "Expected the year mismatch to trigger a confirmation prompt"
    assert FakeWorker.last_kwargs is not None
    assert FakeWorker.last_kwargs["allow_attendance_month_mismatch"] is True


def test_cancel_run_requests_worker_cancellation(
    controller: app_controller.AppController,
    tmp_path: Path,
    monkeypatch,
) -> None:
    monkeypatch.setattr(app_controller, "RunWorker", FakeWorker)
    monkeypatch.setattr(app_controller, "QThread", FakeThread)
    controller.window.values = {
        "source_path": str(tmp_path / "source.xlsx"),
        "target_path": str(tmp_path / "target.xlsx"),
        "attendance_path": "",
        "target_month_name": "aprilie",
        "template_mode_label": TEMPLATE_MODE_PREDEFINED_LABEL,
        "template_sheet_name": "martie",
    }
    controller.start_run()

    controller.cancel_run()

    assert controller._worker.cancel_count == 1
    assert controller.window.status_calls[-1] == ("Se anulează procesarea...", "warning")


def test_cancel_run_without_active_worker_is_safe(controller: app_controller.AppController) -> None:
    # No worker started; must not raise.
    controller.cancel_run()


def test_handle_cancelled_resets_progress(controller: app_controller.AppController) -> None:
    controller._handle_cancelled()

    assert controller.window.progress_calls[-1] == (0, "Anulat")
    assert controller.window.status_calls[-1] == ("Procesarea a fost anulată.", "warning")


def test_handle_success_updates_output_state(controller: app_controller.AppController, tmp_path: Path) -> None:
    result = FakeResult(
        output_file=tmp_path / "done.xlsx",
        warnings=["zi lipsă"],
        differences=["diff"],
    )

    controller._handle_success(result)

    assert controller.last_output_file == result.output_file
    assert controller.window.open_enabled is True
    assert controller.window.progress_calls[-1] == (100, "Completare finalizată")


def test_handle_success_logs_mentiuni_day_coloring(controller: app_controller.AppController, tmp_path: Path) -> None:
    result = FakeResult(output_file=tmp_path / "done.xlsx", warnings=[], differences=[])
    result.attendance_summary = processor.AttendanceApplySummary(
        matched_employees=1,
        co_days_applied=2,
        mentiuni_days_colored=3,
        mentiuni_copied=1,
    )

    controller._handle_success(result)

    assert any("3 zile din mențiuni colorate" in message for message in controller.window.log_messages)
