from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import load_workbook
import pytest

import processor
from tests.conftest import (
    build_attendance_pdf,
    build_attendance_workbook,
    build_predefined_template_workbook,
    build_source_pdf,
    build_source_workbook,
)


def test_parse_full_date_supports_romanian_source_format() -> None:
    assert processor.parse_full_date("01.04.2026") == date(2026, 4, 1)
    assert processor.parse_full_date("not a date") is None


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("16T", 16),
        ("14T(2reprize)", 14),
        ("15(3reprize)", 15),
        ("14T(2reprize), 15T(3reprize)", 14),
    ],
)
def test_transatori_regex_matches_expected_prefix(value: str, expected: int) -> None:
    match = processor.TRANSATORI_RE.search(value)
    assert match is not None
    assert int(float(match.group("count"))) == expected


def test_weighted_transatori_average_uses_reprize_weights() -> None:
    value = processor.weighted_transatori_average(["14T(2reprize)", "16T(1repriza)"])
    assert value == pytest.approx(14.67, rel=0, abs=0.01)


def test_build_value_formula_uses_ovine_when_available() -> None:
    assert processor.build_value_formula(154, 6) == "=154+6/6"
    assert processor.build_value_formula(203, 0) is None


def test_parse_sheet_day_header_supports_excel_style_decimal_labels() -> None:
    assert processor.parse_sheet_day_header(1.04) == 1
    assert processor.parse_sheet_day_header("06.04") == 6
    assert processor.parse_sheet_day_header("foo") is None


def test_cell_indicates_presence_falls_back_to_formula_text() -> None:
    assert processor.cell_indicates_presence(None, "=166+68/6") is True
    assert processor.cell_indicates_presence(0, 0) is False


def test_detect_source_month_from_file_reads_title(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.xlsx"
    build_source_workbook(source_path)

    detected = processor.detect_source_month_from_file(source_path)

    assert detected is not None
    assert detected.month_name == "aprilie"
    assert detected.year == 2026


def test_build_target_day_columns_excludes_only_sundays_and_keeps_holidays() -> None:
    days = processor.build_target_day_columns(2026, 4)

    assert len(days) == 26
    assert date(2026, 4, 12) not in days
    assert date(2026, 4, 13) in days


def test_filter_target_days_to_source_removes_missing_non_sundays() -> None:
    target_days = processor.build_target_day_columns(2026, 4)
    parsed_days = {
        current_day: object()
        for current_day in target_days
        if current_day not in {date(2026, 4, 11), date(2026, 4, 13)}
    }

    included, removed = processor.filter_target_days_to_source(target_days, parsed_days)

    assert len(included) == 24
    assert removed == [date(2026, 4, 11), date(2026, 4, 13)]


def test_build_block_fill_profiles_uses_global_off_fallback_for_blocks_without_local_zero_sample(
    tmp_path: Path,
) -> None:
    preset_path = tmp_path / "preset.xlsx"
    build_predefined_template_workbook(preset_path, include_offless_block=True)

    workbook = load_workbook(preset_path, data_only=False)
    worksheet = workbook["martie"]
    layout = processor.get_sheet_layout(worksheet)
    blocks = processor.detect_sheet_blocks(worksheet, layout)

    profiles = processor.build_block_fill_profiles(worksheet, layout, blocks, allow_global_off_fallback=True)

    second_block_profile = profiles[blocks[1].value_row]
    assert second_block_profile.normal_fills[1].fgColor.rgb == "00FFF9F2"
    assert second_block_profile.off_fills[1].fgColor.rgb == "00FFF2CC"

    workbook.close()


def test_real_preset_has_unsafe_later_header_minus_one_row() -> None:
    preset_path = Path(processor.PREDEFINED_TEMPLATE_WORKBOOK)
    if not preset_path.exists():
        pytest.skip("Preset template not available in this environment")

    workbook = load_workbook(preset_path, data_only=False)
    worksheet = workbook["martie"]
    layout = processor.get_sheet_layout(worksheet)
    blocks = processor.detect_sheet_blocks(worksheet, layout)

    unsafe_pairs = [
        (previous, current)
        for previous, current in zip(blocks, blocks[1:], strict=False)
        if current.header_row - 1 <= previous.diff_row
    ]

    assert unsafe_pairs

    workbook.close()


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("06,07,27,30", (6, 7, 27, 30)),
        ("03 24 (concediu cumulat din ore lipsa)", (3, 24)),
        ("06, 07 24 (cumulate)", (6, 7, 24)),
        ("06;07/24", (6, 7, 24)),
    ],
)
def test_parse_co_days_supports_attendance_formats(value: str, expected: tuple[int, ...]) -> None:
    assert processor.parse_co_days(value) == expected


@pytest.mark.parametrize(
    ("value", "target_month_number", "expected"),
    [
        ("07.03- a plecat la 15:00", 3, (7,)),
        ("06.03- lipsa 11:00-12:30", 3, (6,)),
        ("14.03 plecat ... 30.03 revenit", 3, (14, 30)),
        ("14.04 - altă lună", 3, ()),
        ("07-03 - format cu cratimă", 3, (7,)),
        ("07.03.2026 - format complet", 3, (7,)),
    ],
)
def test_parse_mentiuni_days_reads_only_clear_dates_from_target_month(
    value: str,
    target_month_number: int,
    expected: tuple[int, ...],
) -> None:
    assert processor.parse_mentiuni_days(value, target_month_number) == expected


def test_parse_attendance_workbook_reads_month_co_and_mentiuni(tmp_path: Path) -> None:
    attendance_path = tmp_path / "transatori_detinuti.xlsx"
    build_attendance_workbook(
        attendance_path,
        rows=[("CIOCLEA IOAN", "02,06", "07.04 - a plecat la 15:00")],
    )

    parsed = processor.parse_attendance_workbook(attendance_path)

    assert parsed.month is not None
    assert parsed.month.month_name == "aprilie"
    assert parsed.month.year == 2026
    assert parsed.entries[0].name == "CIOCLEA IOAN"
    assert parsed.entries[0].co_days == (2, 6)
    assert parsed.entries[0].mentiuni_days == (7,)
    assert parsed.entries[0].mentiuni == "07.04 - a plecat la 15:00"


def test_parse_source_pdf_reads_month_and_days(tmp_path: Path) -> None:
    source_path = tmp_path / "situatie.pdf"
    build_source_pdf(source_path)

    detected = processor.detect_source_month_from_file(source_path)
    parsed = processor.parse_source_days(source_path)

    assert detected is not None
    assert detected.month_name == "aprilie"
    assert parsed[date(2026, 4, 1)].bovine_count == 120
    assert parsed[date(2026, 4, 2)].ovine_count == 12
    assert parsed[date(2026, 4, 2)].transatori_average == pytest.approx(14.67, rel=0, abs=0.01)


def test_parse_attendance_pdf_reads_co_n_and_mentiuni(tmp_path: Path) -> None:
    attendance_path = tmp_path / "transatori_detinuti.pdf"
    build_attendance_pdf(
        attendance_path,
        rows=[("CIOCLEA IOAN", "02,06", "03,04", "04.04 - observatie")],
    )

    parsed = processor.parse_attendance_workbook(attendance_path)

    assert parsed.month is not None
    assert parsed.month.month_name == "aprilie"
    assert parsed.entries[0].name == "CIOCLEA IOAN"
    assert parsed.entries[0].co_days == (2, 6)
    assert parsed.entries[0].n_days == (3, 4)
    assert parsed.entries[0].mentiuni_days == (4,)
    assert parsed.entries[0].mentiuni == "04.04 - observatie"
