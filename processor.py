from __future__ import annotations

import calendar
import math
import re
from copy import copy
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path
from typing import Callable, Iterable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from runtime_paths import resource_path

MONTH_NAMES = [
    "ianuarie",
    "februarie",
    "martie",
    "aprilie",
    "mai",
    "iunie",
    "iulie",
    "august",
    "septembrie",
    "octombrie",
    "noiembrie",
    "decembrie",
]
MONTH_NAME_TO_NUMBER = {name: index + 1 for index, name in enumerate(MONTH_NAMES)}
SUMMARY_LABELS = [
    "VACI TRAN ",
    "ZILE",
    "VACA ZI*160",
    "DE PLATIT ",
    "ZILE LUCRATOARE",
    "ZILE IN PLUS",
    "SALAR PE ZI ",
    "BANI IN PLUS ",
    "SALAR BAZA ",
    "TOTAL ",
    "Nume și prenume",
    "PREMIUM SUPLIMENTAR ",
]
TEMPLATE_MODE_PREDEFINED = "predefined"
TEMPLATE_MODE_PREVIOUS_SHEET = "previous_sheet"
PREDEFINED_TEMPLATE_WORKBOOK = resource_path("assets", "salarii_template_preset.xlsx")

FULL_DATE_RE = re.compile(r"^(?P<day>\d{1,2})\.(?P<month>\d{1,2})\.(?P<year>\d{4})$")
FULL_DATE_SEARCH_RE = re.compile(r"(?<!\d)(?P<day>\d{1,2})\.(?P<month>\d{1,2})\.(?P<year>\d{4})(?!\d)")
DATE_INTERVAL_RE = re.compile(
    r"(?<!\d)\d{1,2}\s*[-–]\s*\d{1,2}[.](?P<month>\d{1,2})[.](?P<year>\d{4})(?!\d)"
)
SOURCE_MONTH_RE = re.compile(
    r"\b(?P<month>"
    + "|".join(MONTH_NAMES)
    + r")\b[\s\-_/]*(?P<year>\d{4})",
    re.IGNORECASE,
)
TRANSATORI_RE = re.compile(
    r"(?P<count>\d+(?:[.,]\d+)?)\s*T?(?:\((?P<reprize>\d+)\s*repriz(?:a|e)\))?",
    re.IGNORECASE,
)
CO_DAY_RE = re.compile(r"(?<!\d)(0?[1-9]|[12]\d|3[01])(?!\d)")
MENTIUNI_DATE_RE = re.compile(
    r"(?<!\d)(0?[1-9]|[12]\d|3[01])[.-](0?[1-9]|1[0-2])(?:[.-]\d{4})?(?!\d)"
)
PARENTHETICAL_RE = re.compile(r"\([^)]*\)")
PERSON_NAME_STOP_WORDS = {"de", "din", "si", "a", "al", "ai", "ale", "lui"}
MENTIUNI_HEADER = "Mentiuni"
MENTIUNI_DAY_FILL = PatternFill(fill_type="solid", fgColor="FFF8D6")
TARGET_RATE_VALUE = 166
TARGET_RATE_FILL = PatternFill(fill_type="solid", fgColor="DDEBFF")
WORKDAY_COUNT_FILL = PatternFill(fill_type="solid", fgColor="C6EFCE")


class ProcessorError(RuntimeError):
    pass


@dataclass(frozen=True)
class TargetMonth:
    year: int
    month_number: int
    month_name: str


@dataclass(frozen=True)
class ParsedDay:
    day: date
    bovine_count: float
    ovine_count: float
    value: float | str
    transatori_average: float


@dataclass(frozen=True)
class ProgressEvent:
    percent: float
    message: str


@dataclass
class SheetLayout:
    header_row: int
    day_start_col: int
    day_end_col: int
    summary_start_col: int


@dataclass
class SheetBlock:
    header_row: int
    value_row: int
    transatori_row: int
    per_om_row: int
    norm_row: int
    diff_row: int


@dataclass(frozen=True)
class BlockDayFillProfile:
    normal_fills: tuple
    off_fills: tuple


@dataclass(frozen=True)
class TemplateBuildResult:
    worksheet: Worksheet
    template_source_name: str | None
    source_day_count: int


@dataclass(frozen=True)
class AttendanceEntry:
    source_row: int
    name: str
    co_days: tuple[int, ...]
    mentiuni_days: tuple[int, ...]
    mentiuni: str
    n_days: tuple[int, ...] = ()


@dataclass(frozen=True)
class AttendanceHeader:
    header_row: int
    name_col: int
    co_col: int
    n_col: int | None
    mentiuni_col: int


@dataclass(frozen=True)
class AttendanceWorkbook:
    month: TargetMonth | None
    entries: list[AttendanceEntry]


@dataclass(frozen=True)
class EmployeeBlockMatch:
    entry: AttendanceEntry
    block: SheetBlock
    match_kind: str


@dataclass(frozen=True)
class EmployeeNameCandidate:
    block: SheetBlock
    name: str
    normalized_name: str
    tokens: tuple[str, ...]


@dataclass
class AttendanceApplySummary:
    matched_employees: int = 0
    approximate_matches: int = 0
    co_days_applied: int = 0
    n_days_applied: int = 0
    mentiuni_days_colored: int = 0
    mentiuni_copied: int = 0
    unmatched_names: list[str] = field(default_factory=list)
    ambiguous_names: list[str] = field(default_factory=list)


@dataclass
class RunResult:
    output_file: Path
    mapped_days: int
    updated_blocks: int
    warnings: list[str] = field(default_factory=list)
    differences: list[str] = field(default_factory=list)
    created_sheet_name: str = ""
    template_mode_used: str = ""
    template_source_name: str | None = None
    attendance_summary: AttendanceApplySummary | None = None


def parse_full_date(value) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not isinstance(value, str):
        return None
    match = FULL_DATE_RE.match(value.strip())
    if match is None:
        return None
    return date(
        int(match.group("year")),
        int(match.group("month")),
        int(match.group("day")),
    )


def parse_sheet_day_header(value) -> int | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        text = f"{value:.2f}".rstrip("0").rstrip(".")
    else:
        text = str(value).strip()

    if not text:
        return None
    day_text = text.split(".", 1)[0]
    try:
        day_number = int(day_text)
    except ValueError:
        return None
    return day_number if 1 <= day_number <= 31 else None


def normalize_label(value) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return (
        text.replace("ă", "a")
        .replace("â", "a")
        .replace("î", "i")
        .replace("ș", "s")
        .replace("ş", "s")
        .replace("ț", "t")
        .replace("ţ", "t")
    )


def format_number(value: float) -> str:
    text = f"{value:.12f}".rstrip("0").rstrip(".")
    return text or "0"


def weighted_transatori_average(entries: list[str]) -> float:
    if not entries:
        return 0.0
    
    weighted_total = 0.0
    weight_sum = 0
    for entry in entries:
        match = TRANSATORI_RE.search(entry)
        if match is None:
            continue
        count = float(match.group("count").replace(",", "."))
        # Filter out unreasonable values (transatori should be 5-30, not 331)
        if count > 30:
            continue
        weight = int(match.group("reprize") or "1")
        weighted_total += count * weight
        weight_sum += weight

    if weight_sum == 0:
        return 0.0
    return round(weighted_total / weight_sum, 2)


def build_value_formula(bovine_count: float, ovine_count: float) -> str | None:
    if ovine_count <= 0:
        return None
    return f"={format_number(bovine_count)}+{format_number(ovine_count)}/6"


def cell_indicates_presence(cached_value, raw_value) -> bool:
    def _is_non_zero(value) -> bool:
        if value is None:
            return False
        if isinstance(value, (int, float)):
            return value != 0
        text = str(value).strip()
        if not text:
            return False
        if text.startswith("="):
            return text not in {"=0", "=0.0", "=0.00"}
        try:
            return float(text.replace(",", ".")) != 0
        except ValueError:
            return True

    return _is_non_zero(cached_value) or _is_non_zero(raw_value)


def detect_source_month_from_file(source_path: str | Path) -> TargetMonth | None:
    source_path = Path(source_path)
    if _is_pdf_path(source_path):
        return detect_source_month_from_pdf(source_path)

    workbook = None
    try:
        workbook = load_workbook(source_path, read_only=True, data_only=True)
        sheet = workbook.active
        for row in range(1, min(sheet.max_row, 8) + 1):
            month = _parse_month_from_text(sheet.cell(row, 1).value)
            if month is not None:
                return month

        for row in range(1, min(sheet.max_row, 64) + 1):
            parsed_day = parse_full_date(sheet.cell(row, 1).value)
            if parsed_day is not None:
                return TargetMonth(parsed_day.year, parsed_day.month, MONTH_NAMES[parsed_day.month - 1])
        return None
    except Exception as exc:
        raise ProcessorError("Nu am putut detecta luna din fișierul sursă.") from exc
    finally:
        if workbook is not None:
            workbook.close()


def build_target_day_columns(year: int, month: int) -> list[date]:
    _, last_day = calendar.monthrange(year, month)
    days: list[date] = []
    for day_number in range(1, last_day + 1):
        current = date(year, month, day_number)
        if current.weekday() != 6:
            days.append(current)
    return days


def filter_target_days_to_source(
    target_days: list[date],
    parsed_days: dict[date, ParsedDay],
) -> tuple[list[date], list[date]]:
    parsed_day_set = set(parsed_days)
    included = [current_day for current_day in target_days if current_day in parsed_day_set]
    removed = [current_day for current_day in target_days if current_day not in parsed_day_set]
    return included, removed


def detect_attendance_month_from_file(attendance_path: str | Path) -> TargetMonth | None:
    attendance_path = Path(attendance_path)
    if _is_pdf_path(attendance_path):
        return detect_attendance_month_from_pdf(attendance_path)

    workbook = None
    try:
        workbook = load_workbook(attendance_path, read_only=True, data_only=True)
        return _detect_month_from_sheet(workbook.active)
    except Exception as exc:
        raise ProcessorError("Nu am putut detecta luna din fișierul Transatori+Detinuți.") from exc
    finally:
        if workbook is not None:
            workbook.close()


def parse_attendance_workbook(
    attendance_path: str | Path,
    target_month_number: int | None = None,
    warnings: list[str] | None = None,
) -> AttendanceWorkbook:
    attendance_path = Path(attendance_path)
    if _is_pdf_path(attendance_path):
        return parse_attendance_pdf(
            attendance_path,
            target_month_number=target_month_number,
            warnings=warnings,
        )

    workbook = None
    try:
        workbook = load_workbook(attendance_path, read_only=True, data_only=True)
        sheet = workbook.active
        month = _detect_month_from_sheet(sheet)
        header = _find_attendance_header(sheet)
        if header is None:
            raise ProcessorError(
                "Nu am găsit coloanele Nume și prenume, CO și Mentiuni în fișierul Transatori+Detinuți."
            )

        header_row = header.header_row
        mentiuni_month_number = target_month_number or (month.month_number if month is not None else None)
        entries: list[AttendanceEntry] = []
        for row in range(header_row + 1, sheet.max_row + 1):
            name = _cell_text(sheet.cell(row, header.name_col).value)
            if not name:
                continue
            mentiuni_text = _cell_text(sheet.cell(row, header.mentiuni_col).value)
            entries.append(
                AttendanceEntry(
                    source_row=row,
                    name=name,
                    co_days=parse_co_days(sheet.cell(row, header.co_col).value),
                    mentiuni_days=parse_mentiuni_days(mentiuni_text, mentiuni_month_number),
                    mentiuni=mentiuni_text,
                    n_days=parse_co_days(sheet.cell(row, header.n_col).value) if header.n_col is not None else (),
                )
            )

        if not entries:
            raise ProcessorError("Nu am găsit angajați în fișierul Transatori+Detinuți.")
        return AttendanceWorkbook(month=month, entries=entries)
    except ProcessorError:
        raise
    except Exception as exc:
        raise ProcessorError("Nu am putut citi fișierul Transatori+Detinuți.") from exc
    finally:
        if workbook is not None:
            workbook.close()


def parse_co_days(value) -> tuple[int, ...]:
    text = _cell_text(value)
    if not text:
        return ()

    text = PARENTHETICAL_RE.sub(" ", text)
    text = re.sub(r"[,;/\r\n\t]+", " ", text)
    days: list[int] = []
    seen: set[int] = set()
    for match in CO_DAY_RE.finditer(text):
        day_number = int(match.group(1))
        if day_number not in seen:
            days.append(day_number)
            seen.add(day_number)
    return tuple(days)


def parse_mentiuni_days(value, target_month_number: int | None) -> tuple[int, ...]:
    text = _cell_text(value)
    if not text or target_month_number is None:
        return ()

    days: list[int] = []
    seen: set[int] = set()
    for match in MENTIUNI_DATE_RE.finditer(text):
        day_number = int(match.group(1))
        month_number = int(match.group(2))
        if month_number != target_month_number or day_number in seen:
            continue
        days.append(day_number)
        seen.add(day_number)
    return tuple(days)


def detect_source_month_from_pdf(source_path: str | Path) -> TargetMonth | None:
    try:
        return _detect_month_from_pdf(source_path)
    except ProcessorError:
        raise
    except Exception as exc:
        raise ProcessorError("Nu am putut detecta luna din PDF-ul sursă.") from exc


def detect_attendance_month_from_pdf(attendance_path: str | Path) -> TargetMonth | None:
    try:
        return _detect_month_from_pdf(attendance_path)
    except ProcessorError:
        raise
    except Exception as exc:
        raise ProcessorError("Nu am putut detecta luna din PDF-ul Transatori+Detinuți.") from exc


def parse_source_days_pdf(
    source_path: str | Path,
    warnings: list[str] | None = None,
) -> dict[date, ParsedDay]:
    rows = _extract_pdf_table_rows(source_path)
    if not rows:
        raise ProcessorError("PDF-ul sursă nu conține tabele sau text extractibil.")

    parsed: dict[date, ParsedDay] = {}
    current_day: date | None = None
    current_bovine = 0.0
    current_ovine = 0.0
    current_transatori: list[str] = []

    for row_number, row in enumerate(rows, start=1):
        day_value = _parse_full_date_from_text(_pdf_cell(row, 0))
        if day_value is not None:
            if len(row) < 6:
                _append_warning(warnings, f"PDF sursă: rândul {row_number} are dată, dar nu are toate coloanele.")
                continue
            if current_day is not None:
                parsed[current_day] = finalize_parsed_day(
                    current_day,
                    current_bovine,
                    current_ovine,
                    current_transatori,
                )
            current_day = day_value
            current_bovine = as_number(_pdf_cell(row, 2))
            current_ovine = pick_ovine_count(_pdf_cell(row, 3), _pdf_cell(row, 4))
            current_transatori = _extract_transatori_entries(_pdf_cell(row, 5))
            continue

        if current_day is None:
            continue

        if is_total_row(_pdf_cell(row, 0)):
            break

        current_transatori.extend(_extract_transatori_entries(_pdf_cell(row, 5)))

    if current_day is not None:
        parsed[current_day] = finalize_parsed_day(
            current_day,
            current_bovine,
            current_ovine,
            current_transatori,
        )
    return parsed


def parse_attendance_pdf(
    attendance_path: str | Path,
    target_month_number: int | None = None,
    warnings: list[str] | None = None,
) -> AttendanceWorkbook:
    rows = _extract_pdf_table_rows(attendance_path)
    if not rows:
        raise ProcessorError("PDF-ul Transatori+Detinuți nu conține tabele sau text extractibil.")

    month = detect_attendance_month_from_pdf(attendance_path)
    header = _find_attendance_pdf_header(rows)
    if header is None:
        raise ProcessorError(
            "Nu am găsit coloanele Nume și prenume, CO, N și Mentiuni în PDF-ul Transatori+Detinuți."
        )

    header_row_index, columns = header
    mentiuni_month_number = target_month_number or (month.month_number if month is not None else None)
    entries: list[AttendanceEntry] = []
    for row_index, row in enumerate(rows[header_row_index + 1 :], start=header_row_index + 2):
        if not any(_cell_text(cell) for cell in row):
            continue
        name = _pdf_cell(row, columns["name"])
        if not name:
            if _looks_like_attendance_data_row(row):
                _append_warning(warnings, f"PDF Transatori+Detinuți: rândul {row_index} nu are nume.")
            continue

        mentiuni_text = _pdf_cell(row, columns["mentiuni"])
        entries.append(
            AttendanceEntry(
                source_row=row_index,
                name=name,
                co_days=parse_co_days(_pdf_cell(row, columns["co"])),
                n_days=parse_co_days(_pdf_cell(row, columns["n"])) if columns.get("n") is not None else (),
                mentiuni_days=parse_mentiuni_days(mentiuni_text, mentiuni_month_number),
                mentiuni=mentiuni_text,
            )
        )

    if not entries:
        raise ProcessorError("Nu am găsit angajați în PDF-ul Transatori+Detinuți.")
    return AttendanceWorkbook(month=month, entries=entries)


def run_fill(
    source_path: str | Path,
    target_path: str | Path,
    output_dir: str | Path,
    template_mode: str,
    target_month_name: str,
    template_sheet_name: str | None = None,
    attendance_path: str | Path | None = None,
    allow_attendance_month_mismatch: bool = False,
    progress_callback: Callable[[ProgressEvent], None] | None = None,
) -> RunResult:
    source_path = Path(source_path)
    target_path = Path(target_path)
    output_dir = Path(output_dir)
    attendance_path = Path(attendance_path) if attendance_path else None

    if not source_path.exists():
        raise ProcessorError("Fișierul sursă nu există.")
    if not target_path.exists():
        raise ProcessorError("Workbook-ul de salarii nu există.")
    if attendance_path is not None and not attendance_path.exists():
        raise ProcessorError("Fișierul Transatori+Detinuți nu există.")
    if template_mode not in {TEMPLATE_MODE_PREDEFINED, TEMPLATE_MODE_PREVIOUS_SHEET}:
        raise ProcessorError("Modul de template selectat nu este valid.")

    target_month_name = normalize_label(target_month_name)
    if target_month_name not in MONTH_NAME_TO_NUMBER:
        raise ProcessorError("Luna țintă selectată nu este validă.")

    _emit(progress_callback, 5, "Analizez fișierul sursă")
    source_month = detect_source_month_from_file(source_path)
    if source_month is not None and source_month.month_name != target_month_name:
        raise ProcessorError(
            f"Luna din sursă este {source_month.month_name}, dar ai selectat {target_month_name}."
        )

    warnings: list[str] = []
    parsed_days = parse_source_days(source_path, warnings=warnings)
    if not parsed_days:
        raise ProcessorError("Nu am găsit zile valide în fișierul sursă.")

    year = source_month.year if source_month is not None else min(parsed_days).year
    month_number = MONTH_NAME_TO_NUMBER[target_month_name]
    all_target_days = build_target_day_columns(year, month_number)
    target_days, removed_days = filter_target_days_to_source(all_target_days, parsed_days)
    if not target_days:
        raise ProcessorError("Nu am găsit zile lucrătoare valide din luna țintă în fișierul sursă.")

    _emit(progress_callback, 20, "Deschid workbook-ul de salarii")
    workbook = load_workbook(target_path, data_only=False)
    template_workbook = None
    try:
        if template_mode == TEMPLATE_MODE_PREDEFINED:
            _emit(progress_callback, 30, "Construiesc foaia nouă din template-ul predefinit")
            template_workbook = load_workbook(PREDEFINED_TEMPLATE_WORKBOOK, data_only=False)
            build_result = create_sheet_from_predefined_template(
                workbook=workbook,
                template_workbook=template_workbook,
                target_sheet_name=target_month_name,
            )
        else:
            _emit(progress_callback, 30, "Construiesc foaia nouă din sheet-ul anterior")
            build_result = create_sheet_from_previous_sheet(
                workbook=workbook,
                target_sheet_name=target_month_name,
                template_sheet_name=template_sheet_name,
            )

        source_layout = get_sheet_layout(build_result.worksheet)
        source_blocks = detect_sheet_blocks(build_result.worksheet, source_layout)
        template_off_fills = _find_global_off_fills(build_result.worksheet, source_layout, source_blocks)

        _emit(progress_callback, 45, "Rearanjez layout-ul pentru luna țintă")
        rebuild_generated_month_layout(build_result.worksheet, target_days)

        layout = get_sheet_layout(build_result.worksheet)
        blocks = detect_sheet_blocks(build_result.worksheet, layout)
        if not blocks:
            raise ProcessorError("Nu am putut detecta blocurile de completare din foaia generată.")

        warnings.extend(build_removed_day_warnings(removed_days))
        differences: list[str] = []

        _emit(progress_callback, 65, "Completez zilele și media ponderată a transatorilor")
        fill_generated_sheet(
            worksheet=build_result.worksheet,
            layout=layout,
            blocks=blocks,
            target_days=target_days,
            parsed_days=parsed_days,
            differences=differences,
            template_mode=template_mode,
            fallback_off_fills=template_off_fills,
        )

        attendance_summary = None
        if attendance_path is not None:
            _emit(progress_callback, 74, "Aplic CO, N și mențiunile din Transatori+Detinuți")
            attendance_data = parse_attendance_workbook(
                attendance_path,
                target_month_number=month_number,
                warnings=warnings,
            )
            if attendance_data.month is not None:
                attendance_month_matches = (
                    attendance_data.month.year == year
                    and attendance_data.month.month_number == month_number
                )
                if not attendance_month_matches and not allow_attendance_month_mismatch:
                    raise ProcessorError(
                        "Luna din Transatori+Detinuți este "
                        f"{attendance_data.month.month_name} {attendance_data.month.year}, "
                        f"dar luna țintă este {target_month_name} {year}."
                    )
            attendance_summary = apply_attendance_adjustments(
                worksheet=build_result.worksheet,
                layout=layout,
                blocks=blocks,
                target_days=target_days,
                attendance_data=attendance_data,
                template_mode=template_mode,
                fallback_off_fills=template_off_fills,
            )
            layout = get_sheet_layout(build_result.worksheet)
            blocks = detect_sheet_blocks(build_result.worksheet, layout)

        _emit(progress_callback, 80, "Rescriu formulele de sumar și recalcul")
        rewrite_summary_formulas(build_result.worksheet, layout, blocks, len(target_days))
        enable_full_recalculation(workbook)

        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = unique_output_path(
            output_dir,
            f"{target_path.stem} - {target_month_name} completat{target_path.suffix}",
        )

        _emit(progress_callback, 95, "Salvez workbook-ul final")
        workbook.save(output_file)
        return RunResult(
            output_file=output_file,
            mapped_days=len(parsed_days),
            updated_blocks=len(blocks),
            warnings=warnings,
            differences=differences,
            created_sheet_name=target_month_name,
            template_mode_used=template_mode,
            template_source_name=build_result.template_source_name,
            attendance_summary=attendance_summary,
        )
    except FileNotFoundError as exc:
        raise ProcessorError(
            f"Nu găsesc workbook-ul preset pentru template: {PREDEFINED_TEMPLATE_WORKBOOK}"
        ) from exc
    finally:
        if template_workbook is not None:
            template_workbook.close()
        workbook.close()


def parse_source_days(source_path: str | Path, warnings: list[str] | None = None) -> dict[date, ParsedDay]:
    source_path = Path(source_path)
    if _is_pdf_path(source_path):
        return parse_source_days_pdf(source_path, warnings=warnings)

    workbook = load_workbook(source_path, data_only=True, read_only=True)
    sheet = workbook.active
    parsed: dict[date, ParsedDay] = {}
    current_day: date | None = None
    current_bovine = 0.0
    current_ovine = 0.0
    current_transatori: list[str] = []

    try:
        for row in sheet.iter_rows(min_row=1, values_only=True):
            day_value = parse_full_date(row[0])
            if day_value is not None:
                if current_day is not None:
                    parsed[current_day] = finalize_parsed_day(
                        current_day,
                        current_bovine,
                        current_ovine,
                        current_transatori,
                    )
                current_day = day_value
                current_bovine = as_number(row[2])
                current_ovine = pick_ovine_count(row[3], row[4])
                current_transatori = []
                if row[5]:
                    entry_str = str(row[5]).strip()
                    if TRANSATORI_RE.search(entry_str):
                        current_transatori.append(entry_str)
                continue

            if current_day is None:
                continue

            if is_total_row(row[0]):
                break

            if row[5]:
                entry_str = str(row[5]).strip()
                if TRANSATORI_RE.search(entry_str):
                    current_transatori.append(entry_str)

        if current_day is not None:
            parsed[current_day] = finalize_parsed_day(
                current_day,
                current_bovine,
                current_ovine,
                current_transatori,
            )
    finally:
        workbook.close()
    return parsed


def finalize_parsed_day(
    current_day: date,
    bovine_count: float,
    ovine_count: float,
    transatori_entries: list[str],
) -> ParsedDay:
    formula = build_value_formula(bovine_count, ovine_count)
    return ParsedDay(
        day=current_day,
        bovine_count=bovine_count,
        ovine_count=ovine_count,
        value=formula if formula is not None else bovine_count,
        transatori_average=weighted_transatori_average(transatori_entries),
    )


def create_sheet_from_predefined_template(
    workbook: Workbook,
    template_workbook: Workbook,
    target_sheet_name: str,
) -> TemplateBuildResult:
    template_sheet = pick_template_sheet(template_workbook)
    if target_sheet_name in workbook.sheetnames:
        del workbook[target_sheet_name]

    insert_index = first_month_sheet_index(workbook)
    created = clone_sheet_between_workbooks(
        source_sheet=template_sheet,
        target_workbook=workbook,
        new_title=target_sheet_name,
        insert_index=insert_index,
    )
    layout = get_sheet_layout(created)
    return TemplateBuildResult(
        worksheet=created,
        template_source_name=template_sheet.title,
        source_day_count=layout.day_end_col - layout.day_start_col + 1,
    )


def create_sheet_from_previous_sheet(
    workbook: Workbook,
    target_sheet_name: str,
    template_sheet_name: str | None,
) -> TemplateBuildResult:
    if not template_sheet_name:
        raise ProcessorError("Trebuie să alegi un sheet template pentru modul de copiere din sheet anterior.")
    if template_sheet_name not in workbook.sheetnames:
        raise ProcessorError(f"Sheet-ul template '{template_sheet_name}' nu există în workbook.")
    if normalize_label(template_sheet_name) == normalize_label(target_sheet_name):
        raise ProcessorError("Sheet-ul template nu poate fi aceeași lună cu foaia care trebuie generată.")

    if target_sheet_name in workbook.sheetnames:
        del workbook[target_sheet_name]

    source_sheet = workbook[template_sheet_name]
    created = workbook.copy_worksheet(source_sheet)
    created.title = target_sheet_name
    created.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale

    insert_index = workbook.sheetnames.index(template_sheet_name)
    workbook._sheets.remove(created)
    workbook._sheets.insert(insert_index, created)

    layout = get_sheet_layout(created)
    return TemplateBuildResult(
        worksheet=created,
        template_source_name=template_sheet_name,
        source_day_count=layout.day_end_col - layout.day_start_col + 1,
    )


def rebuild_generated_month_layout(worksheet: Worksheet, target_days: list[date]) -> None:
    layout = get_sheet_layout(worksheet)
    current_day_count = layout.day_end_col - layout.day_start_col + 1
    target_day_count = len(target_days)
    delta = target_day_count - current_day_count

    if delta > 0:
        worksheet.insert_cols(layout.summary_start_col, delta)
        for offset in range(delta):
            clone_column_format(
                worksheet,
                source_col=layout.day_end_col,
                target_col=layout.day_end_col + 1 + offset,
            )
    elif delta < 0:
        worksheet.delete_cols(layout.day_start_col + target_day_count, -delta)

    layout = get_sheet_layout(worksheet)
    blocks = detect_sheet_blocks(worksheet, layout)
    for block in blocks:
        for offset, current_day in enumerate(target_days):
            worksheet.cell(block.header_row, layout.day_start_col + offset).value = build_sheet_day_header_value(current_day)
        for offset, label in enumerate(SUMMARY_LABELS):
            worksheet.cell(block.header_row, layout.summary_start_col + offset).value = label

    worksheet.title = MONTH_NAMES[target_days[0].month - 1]


def rewrite_summary_formulas(
    worksheet: Worksheet,
    layout: SheetLayout,
    blocks: list[SheetBlock],
    workday_count: int,
) -> None:
    if not blocks:
        return

    first_day_col = get_column_letter(layout.day_start_col)
    last_day_col = get_column_letter(layout.day_end_col)
    target_rate_col = layout.summary_start_col + 2
    target_rate_row = blocks[0].header_row - 1
    target_rate_ref = f"${get_column_letter(target_rate_col)}${target_rate_row}"
    _write_target_rate_cell(worksheet, target_rate_row, target_rate_col)

    workday_count_col = layout.summary_start_col + 4
    workday_count_row = blocks[0].header_row - 1
    workday_count_ref = f"${get_column_letter(workday_count_col)}${workday_count_row}"
    _write_workday_count_cell(worksheet, workday_count_row, workday_count_col, workday_count)

    for block in blocks:
        for day_col in range(layout.day_start_col, layout.day_end_col + 1):
            col_letter = get_column_letter(day_col)
            value_ref = f"{col_letter}{block.value_row}"
            trans_ref = f"{col_letter}{block.transatori_row}"
            worksheet.cell(block.per_om_row, day_col).value = f'=IF({trans_ref}>0,{value_ref}/{trans_ref},0)'
            worksheet.cell(block.norm_row, day_col).value = f'=IF({trans_ref}>0,11.5,0)'
            worksheet.cell(block.diff_row, day_col).value = (
                f'=IF({trans_ref}>0,{col_letter}{block.per_om_row}-{col_letter}{block.norm_row},0)'
            )

        sum_col = layout.summary_start_col
        zile_col = sum_col + 1
        target_col = sum_col + 2
        de_platit_col = sum_col + 3
        zile_lucratoare_col = sum_col + 4
        zile_in_plus_col = sum_col + 5
        salar_pe_zi_col = sum_col + 6
        bani_in_plus_col = sum_col + 7
        salar_baza_col = sum_col + 8
        total_col = sum_col + 9
        nume_col = sum_col + 10
        premium_col = sum_col + 11

        sum_ref = f"{get_column_letter(sum_col)}{block.value_row}"
        zile_ref = f"{get_column_letter(zile_col)}{block.value_row}"
        target_ref = f"{get_column_letter(target_col)}{block.value_row}"
        de_platit_ref = f"{get_column_letter(de_platit_col)}{block.value_row}"
        zile_lucratoare_ref = f"{get_column_letter(zile_lucratoare_col)}{block.value_row}"
        zile_in_plus_ref = f"{get_column_letter(zile_in_plus_col)}{block.value_row}"
        salar_pe_zi_ref = f"{get_column_letter(salar_pe_zi_col)}{block.value_row}"
        bani_in_plus_ref = f"{get_column_letter(bani_in_plus_col)}{block.value_row}"
        salar_baza_ref = f"{get_column_letter(salar_baza_col)}{block.value_row}"

        preserved_salar_baza = worksheet.cell(block.value_row, salar_baza_col).value
        preserved_nume = worksheet.cell(block.value_row, nume_col).value
        preserved_premium = worksheet.cell(block.value_row, premium_col).value

        _clear_summary_row(
            worksheet,
            block.value_row,
            sum_col,
            premium_col,
            preserve_cols={salar_baza_col, nume_col, premium_col},
        )
        for row in (block.per_om_row, block.norm_row, block.diff_row):
            _clear_summary_row(worksheet, row, sum_col, premium_col)

        worksheet.cell(block.value_row, sum_col).value = f"=SUM({first_day_col}{block.value_row}:{last_day_col}{block.value_row})"
        worksheet.cell(block.value_row, zile_col).value = (
            f'=COUNTIF({first_day_col}{block.transatori_row}:{last_day_col}{block.transatori_row},">0")'
        )
        worksheet.cell(block.value_row, target_col).value = f"={zile_ref}*{target_rate_ref}"
        worksheet.cell(block.value_row, de_platit_col).value = f"={sum_ref}-{target_ref}"
        worksheet.cell(block.value_row, zile_lucratoare_col).value = f"={workday_count_ref}"
        worksheet.cell(block.value_row, zile_in_plus_col).value = f"={de_platit_ref}/175"
        worksheet.cell(block.value_row, salar_pe_zi_col).value = f"={salar_baza_ref}/{zile_lucratoare_ref}"
        worksheet.cell(block.value_row, bani_in_plus_col).value = f"={salar_pe_zi_ref}*{zile_in_plus_ref}"
        worksheet.cell(block.value_row, total_col).value = f"={salar_baza_ref}+{bani_in_plus_ref}"
        worksheet.cell(block.value_row, salar_baza_col).value = preserved_salar_baza
        worksheet.cell(block.value_row, premium_col).value = preserved_premium

        if preserved_nume not in {None, ""}:
            worksheet.cell(block.value_row, nume_col).value = preserved_nume
        else:
            worksheet.cell(block.value_row, nume_col).value = worksheet.cell(block.value_row, 2).value

        worksheet.cell(block.per_om_row, sum_col).value = f"=SUM({first_day_col}{block.per_om_row}:{last_day_col}{block.per_om_row})"
        worksheet.cell(block.per_om_row, target_col).value = "cost /vaca"
        worksheet.cell(block.per_om_row, zile_lucratoare_col).value = (
            f"={get_column_letter(salar_pe_zi_col)}{block.value_row}/11.5"
        )

        worksheet.cell(block.norm_row, sum_col).value = f"=SUM({first_day_col}{block.norm_row}:{last_day_col}{block.norm_row})"
        worksheet.cell(block.norm_row, target_col).value = "bani  in plus"
        worksheet.cell(block.norm_row, zile_lucratoare_col).value = (
            f"={get_column_letter(sum_col)}{block.diff_row}*{get_column_letter(zile_lucratoare_col)}{block.per_om_row}"
        )

        worksheet.cell(block.diff_row, sum_col).value = f"=SUM({first_day_col}{block.diff_row}:{last_day_col}{block.diff_row})"
        worksheet.cell(block.diff_row, zile_in_plus_col).value = (
            f"={get_column_letter(de_platit_col)}{block.diff_row}/175"
        )


def _write_target_rate_cell(worksheet: Worksheet, row: int, col: int) -> None:
    cell = worksheet.cell(row, col)
    cell.value = TARGET_RATE_VALUE
    cell.fill = copy(TARGET_RATE_FILL)

    font = copy(cell.font)
    font.bold = True
    cell.font = font

    alignment = copy(cell.alignment)
    alignment.horizontal = "center"
    alignment.vertical = "center"
    cell.alignment = alignment


def _write_workday_count_cell(worksheet: Worksheet, row: int, col: int, workday_count: int) -> None:
    cell = worksheet.cell(row, col)
    cell.value = workday_count
    cell.fill = copy(WORKDAY_COUNT_FILL)

    font = copy(cell.font)
    font.bold = True
    cell.font = font

    alignment = copy(cell.alignment)
    alignment.horizontal = "center"
    alignment.vertical = "center"
    cell.alignment = alignment


def _clear_summary_row(
    worksheet: Worksheet,
    row: int,
    start_col: int,
    end_col: int,
    preserve_cols: set[int] | None = None,
) -> None:
    preserve_cols = preserve_cols or set()
    for col in range(start_col, min(end_col, worksheet.max_column) + 1):
        if col in preserve_cols:
            continue
        worksheet.cell(row, col).value = None


def fill_generated_sheet(
    worksheet: Worksheet,
    layout: SheetLayout,
    blocks: list[SheetBlock],
    target_days: list[date],
    parsed_days: dict[date, ParsedDay],
    differences: list[str],
    template_mode: str,
    fallback_off_fills: tuple | None = None,
) -> None:
    uses_predefined_template = template_mode == TEMPLATE_MODE_PREDEFINED
    fill_profiles = build_block_fill_profiles(
        worksheet,
        layout,
        blocks,
        allow_global_off_fallback=uses_predefined_template,
        fallback_off_fills=fallback_off_fills,
    )
    for block in blocks:
        profile = fill_profiles.get(block.value_row)
        for offset, current_day in enumerate(target_days):
            col = layout.day_start_col + offset
            parsed = parsed_days.get(current_day)
            has_source_day = parsed is not None
            new_value = parsed.value if has_source_day else 0
            new_transatori = parsed.transatori_average if has_source_day else 0

            old_value = worksheet.cell(block.value_row, col).value
            old_transatori = worksheet.cell(block.transatori_row, col).value
            if old_value != new_value:
                differences.append(
                    f"{worksheet.title}:{get_column_letter(col)}{block.value_row} {old_value!r} -> {new_value!r}"
                )
            if old_transatori != new_transatori:
                differences.append(
                    f"{worksheet.title}:{get_column_letter(col)}{block.transatori_row} {old_transatori!r} -> {new_transatori!r}"
                )

            worksheet.cell(block.value_row, col).value = new_value
            worksheet.cell(block.transatori_row, col).value = new_transatori
            use_off_style = (not has_source_day) or (
                uses_predefined_template and _is_zero_result_pair(new_value, new_transatori)
            )
            apply_block_fill_profile(worksheet, block, col, profile, use_off_style=use_off_style)


def apply_attendance_adjustments(
    worksheet: Worksheet,
    layout: SheetLayout,
    blocks: list[SheetBlock],
    target_days: list[date],
    attendance_data: AttendanceWorkbook,
    template_mode: str,
    fallback_off_fills: tuple | None = None,
) -> AttendanceApplySummary:
    summary = AttendanceApplySummary()
    mentiuni_col = ensure_mentiuni_column(worksheet, layout, blocks)
    fill_profiles = build_block_fill_profiles(
        worksheet,
        layout,
        blocks,
        allow_global_off_fallback=True,
        fallback_off_fills=fallback_off_fills,
    )
    matches = match_attendance_entries_to_blocks(worksheet, blocks, attendance_data.entries, summary)
    day_col_by_number = {
        current_day.day: layout.day_start_col + offset
        for offset, current_day in enumerate(target_days)
    }

    for match in matches:
        block = match.block
        profile = fill_profiles.get(block.value_row)
        absence_day_set: set[int] = set()
        absence_days: list[tuple[int, str]] = []
        for day_number in match.entry.co_days:
            if day_number in absence_day_set:
                continue
            absence_day_set.add(day_number)
            absence_days.append((day_number, "co"))
        for day_number in match.entry.n_days:
            if day_number in absence_day_set:
                continue
            absence_day_set.add(day_number)
            absence_days.append((day_number, "n"))

        for day_number, absence_kind in absence_days:
            col = day_col_by_number.get(day_number)
            if col is None:
                continue
            value_cell = worksheet.cell(block.value_row, col)
            transatori_cell = worksheet.cell(block.transatori_row, col)
            if _is_zero_result_pair(value_cell.value, transatori_cell.value):
                continue
            value_cell.value = 0
            transatori_cell.value = 0
            apply_block_fill_profile(worksheet, block, col, profile, use_off_style=True)
            if absence_kind == "n":
                summary.n_days_applied += 1
            else:
                summary.co_days_applied += 1

        for day_number in match.entry.mentiuni_days:
            if day_number in absence_day_set:
                continue
            col = day_col_by_number.get(day_number)
            if col is None:
                continue
            apply_mentiuni_day_fill(worksheet, block, col)
            summary.mentiuni_days_colored += 1

        if match.entry.mentiuni:
            note_cell = worksheet.cell(block.value_row, mentiuni_col)
            note_cell.value = match.entry.mentiuni
            if profile is not None:
                note_cell.fill = copy(profile.off_fills[1])
            note_alignment = copy(note_cell.alignment)
            note_alignment.wrap_text = True
            note_cell.alignment = note_alignment
            _adjust_row_height_for_note(worksheet, block.value_row, match.entry.mentiuni)
            summary.mentiuni_copied += 1

    return summary


def apply_mentiuni_day_fill(worksheet: Worksheet, block: SheetBlock, col: int) -> None:
    for row in _block_day_rows(block):
        worksheet.cell(row, col).fill = copy(MENTIUNI_DAY_FILL)


def ensure_mentiuni_column(
    worksheet: Worksheet,
    layout: SheetLayout,
    blocks: list[SheetBlock],
) -> int:
    premium_col = layout.summary_start_col + len(SUMMARY_LABELS) - 1
    mentiuni_col = premium_col + 1
    if not _column_is_empty(worksheet, mentiuni_col):
        worksheet.insert_cols(mentiuni_col)

    clone_column_format(worksheet, premium_col, mentiuni_col)
    premium_letter = get_column_letter(premium_col)
    mentiuni_letter = get_column_letter(mentiuni_col)
    source_width = worksheet.column_dimensions[premium_letter].width
    if source_width is not None:
        worksheet.column_dimensions[mentiuni_letter].width = source_width

    for block in blocks:
        for row in (
            block.header_row,
            block.value_row,
            block.transatori_row,
            block.per_om_row,
            block.norm_row,
            block.diff_row,
        ):
            source_cell = worksheet.cell(row, premium_col)
            target_cell = worksheet.cell(row, mentiuni_col)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)

        worksheet.cell(block.header_row, mentiuni_col).value = MENTIUNI_HEADER
        header_alignment = copy(worksheet.cell(block.header_row, mentiuni_col).alignment)
        header_alignment.wrap_text = True
        worksheet.cell(block.header_row, mentiuni_col).alignment = header_alignment

    return mentiuni_col


def match_attendance_entries_to_blocks(
    worksheet: Worksheet,
    blocks: list[SheetBlock],
    entries: list[AttendanceEntry],
    summary: AttendanceApplySummary,
) -> list[EmployeeBlockMatch]:
    candidates = _build_employee_name_candidates(worksheet, blocks)
    exact_by_name: dict[str, list[EmployeeNameCandidate]] = {}
    for candidate in candidates:
        exact_by_name.setdefault(candidate.normalized_name, []).append(candidate)

    matches: list[EmployeeBlockMatch] = []
    for entry in entries:
        normalized_entry = normalize_person_name(entry.name)
        exact_matches = exact_by_name.get(normalized_entry, [])
        if len(exact_matches) == 1:
            matches.append(EmployeeBlockMatch(entry=entry, block=exact_matches[0].block, match_kind="exact"))
            summary.matched_employees += 1
            continue
        if len(exact_matches) > 1:
            summary.ambiguous_names.append(_format_ambiguous_name(entry.name, exact_matches))
            continue

        approximate_matches = _find_approximate_name_matches(entry.name, candidates)
        if len(approximate_matches) == 1:
            matches.append(
                EmployeeBlockMatch(
                    entry=entry,
                    block=approximate_matches[0].block,
                    match_kind="approximate",
                )
            )
            summary.matched_employees += 1
            summary.approximate_matches += 1
            continue
        if len(approximate_matches) > 1:
            summary.ambiguous_names.append(_format_ambiguous_name(entry.name, approximate_matches))
            continue

        summary.unmatched_names.append(entry.name)

    return matches


def normalize_person_name(value) -> str:
    normalized = normalize_label(value)
    normalized = re.sub(r"[^a-z0-9]+", " ", normalized)
    return re.sub(r"\s+", " ", normalized).strip()


def meaningful_name_tokens(value) -> tuple[str, ...]:
    tokens = []
    for token in normalize_person_name(value).split():
        if len(token) >= 3 and token not in PERSON_NAME_STOP_WORDS:
            tokens.append(token)
    return tuple(tokens)


def _build_employee_name_candidates(
    worksheet: Worksheet,
    blocks: list[SheetBlock],
) -> list[EmployeeNameCandidate]:
    candidates: list[EmployeeNameCandidate] = []
    for block in blocks:
        name = _cell_text(worksheet.cell(block.value_row, 2).value)
        if not name:
            continue
        candidates.append(
            EmployeeNameCandidate(
                block=block,
                name=name,
                normalized_name=normalize_person_name(name),
                tokens=meaningful_name_tokens(name),
            )
        )
    return candidates


def _find_approximate_name_matches(
    source_name: str,
    candidates: list[EmployeeNameCandidate],
) -> list[EmployeeNameCandidate]:
    source_tokens = meaningful_name_tokens(source_name)
    source_token_set = set(source_tokens)
    source_first_token = source_tokens[0] if source_tokens else ""
    matches: list[EmployeeNameCandidate] = []

    for candidate in candidates:
        candidate_token_set = set(candidate.tokens)
        overlap = source_token_set & candidate_token_set
        if len(overlap) >= 2:
            matches.append(candidate)
            continue
        candidate_first_token = candidate.tokens[0] if candidate.tokens else ""
        if (
            len(overlap) >= 1
            and source_first_token
            and candidate_first_token
            and _edit_distance_at_most_one(source_first_token, candidate_first_token)
        ):
            matches.append(candidate)

    return matches


def _edit_distance_at_most_one(left: str, right: str) -> bool:
    if left == right:
        return True
    if abs(len(left) - len(right)) > 1:
        return False

    if len(left) == len(right):
        differences = sum(1 for left_char, right_char in zip(left, right) if left_char != right_char)
        return differences <= 1

    shorter, longer = (left, right) if len(left) < len(right) else (right, left)
    short_index = 0
    long_index = 0
    skipped = False
    while short_index < len(shorter) and long_index < len(longer):
        if shorter[short_index] == longer[long_index]:
            short_index += 1
            long_index += 1
            continue
        if skipped:
            return False
        skipped = True
        long_index += 1
    return True


def _format_ambiguous_name(source_name: str, candidates: list[EmployeeNameCandidate]) -> str:
    candidate_names = ", ".join(candidate.name for candidate in candidates)
    return f"{source_name}: {candidate_names}"


def _column_is_empty(worksheet: Worksheet, col: int) -> bool:
    if col > worksheet.max_column:
        return True
    for row in range(1, worksheet.max_row + 1):
        value = worksheet.cell(row, col).value
        if value not in {None, ""}:
            return False
    return True


def _adjust_row_height_for_note(worksheet: Worksheet, row: int, note: str) -> None:
    existing_height = worksheet.row_dimensions[row].height or 15
    note_height = min(90, 18 * max(1, math.ceil(len(note) / 35)))
    worksheet.row_dimensions[row].height = max(existing_height, note_height)


def build_block_fill_profiles(
    worksheet: Worksheet,
    layout: SheetLayout,
    blocks: list[SheetBlock],
    allow_global_off_fallback: bool = False,
    fallback_off_fills: tuple | None = None,
) -> dict[int, BlockDayFillProfile]:
    global_off_fills = _find_global_off_fills(worksheet, layout, blocks) if allow_global_off_fallback else None
    if global_off_fills is None and fallback_off_fills is not None:
        global_off_fills = tuple(copy(fill) for fill in fallback_off_fills)
    profiles: dict[int, BlockDayFillProfile] = {}
    for block in blocks:
        normal_col = None
        off_col = None
        for col in range(layout.day_start_col, layout.day_end_col + 1):
            value_cell = worksheet.cell(block.value_row, col).value
            trans_cell = worksheet.cell(block.transatori_row, col).value
            if normal_col is None and cell_indicates_presence(value_cell, trans_cell):
                normal_col = col
            if off_col is None and not cell_indicates_presence(value_cell, trans_cell):
                off_col = col
            if normal_col is not None and off_col is not None:
                break

        if normal_col is None:
            normal_col = layout.day_start_col

        normal_fills = _capture_block_fills(worksheet, block, normal_col)
        if off_col is not None:
            off_fills = _capture_block_fills(worksheet, block, off_col)
        elif global_off_fills is not None:
            off_fills = tuple(copy(fill) for fill in global_off_fills)
        else:
            off_fills = tuple(copy(fill) for fill in normal_fills)

        profiles[block.value_row] = BlockDayFillProfile(
            normal_fills=normal_fills,
            off_fills=off_fills,
        )
    return profiles


def apply_block_fill_profile(
    worksheet: Worksheet,
    block: SheetBlock,
    col: int,
    profile: BlockDayFillProfile | None,
    use_off_style: bool,
) -> None:
    if profile is None:
        return
    fills = profile.off_fills if use_off_style else profile.normal_fills
    for row, fill in zip(_block_day_rows(block), fills, strict=False):
        worksheet.cell(row, col).fill = copy(fill)


def _capture_block_fills(worksheet: Worksheet, block: SheetBlock, col: int) -> tuple:
    return tuple(copy(worksheet.cell(row, col).fill) for row in _block_day_rows(block))


def _block_day_rows(block: SheetBlock) -> tuple[int, ...]:
    return (
        block.header_row,
        block.value_row,
        block.transatori_row,
        block.per_om_row,
        block.norm_row,
        block.diff_row,
    )


def _find_global_off_fills(
    worksheet: Worksheet,
    layout: SheetLayout,
    blocks: list[SheetBlock],
) -> tuple | None:
    for block in blocks:
        for col in range(layout.day_start_col, layout.day_end_col + 1):
            value_cell = worksheet.cell(block.value_row, col).value
            trans_cell = worksheet.cell(block.transatori_row, col).value
            if not cell_indicates_presence(value_cell, trans_cell):
                return _capture_block_fills(worksheet, block, col)
    return None


def _is_zero_result_pair(value, transatori) -> bool:
    return isinstance(value, (int, float)) and value == 0 and isinstance(transatori, (int, float)) and transatori == 0


def get_sheet_layout(worksheet: Worksheet) -> SheetLayout:
    blocks = detect_sheet_blocks(worksheet, None)
    if not blocks:
        raise ProcessorError("Nu am găsit un bloc valid în sheet-ul template.")

    header_row = blocks[0].header_row
    day_start_col = 3
    day_end_col = 0
    summary_start_col = 0

    for col in range(day_start_col, worksheet.max_column + 1):
        value = worksheet.cell(header_row, col).value
        if parse_sheet_day_header(value) is not None:
            day_end_col = col
            continue
        if day_end_col != 0:
            summary_start_col = col
            break

    if day_end_col == 0 or summary_start_col == 0:
        raise ProcessorError("Nu am putut identifica corect coloanele de zile și sumar.")

    return SheetLayout(
        header_row=header_row,
        day_start_col=day_start_col,
        day_end_col=day_end_col,
        summary_start_col=summary_start_col,
    )


def detect_sheet_blocks(worksheet: Worksheet, layout: SheetLayout | None) -> list[SheetBlock]:
    blocks: list[SheetBlock] = []
    for row in range(1, worksheet.max_row - 1):
        if normalize_label(worksheet.cell(row, 2).value) != "vaci pe om":
            continue
        if normalize_label(worksheet.cell(row + 1, 2).value) != "vaci 160 in 14":
            continue
        if not normalize_label(worksheet.cell(row + 2, 2).value).startswith("difer"):
            continue

        header_row = row - 3
        value_row = row - 2
        transatori_row = row - 1

        blocks.append(
            SheetBlock(
                header_row=header_row,
                value_row=value_row,
                transatori_row=transatori_row,
                per_om_row=row,
                norm_row=row + 1,
                diff_row=row + 2,
            )
        )
    return blocks


def build_removed_day_warnings(removed_days: list[date]) -> list[str]:
    if not removed_days:
        return []
    formatted_days = ", ".join(current_day.strftime("%d.%m.%Y") for current_day in removed_days)
    return [f"Zile eliminate din tabel pentru că lipsesc din sursă: {formatted_days}"]


def pick_template_sheet(workbook: Workbook) -> Worksheet:
    for month_name in MONTH_NAMES:
        if month_name in workbook.sheetnames:
            return workbook[month_name]
    return workbook.active


def first_month_sheet_index(workbook: Workbook) -> int:
    for index, name in enumerate(workbook.sheetnames):
        if normalize_label(name) in MONTH_NAME_TO_NUMBER:
            return index
    return 0


def clone_sheet_between_workbooks(
    source_sheet: Worksheet,
    target_workbook: Workbook,
    new_title: str,
    insert_index: int,
) -> Worksheet:
    target_sheet = target_workbook.create_sheet(title=new_title, index=insert_index)
    target_sheet.sheet_format = copy(source_sheet.sheet_format)
    target_sheet.sheet_properties = copy(source_sheet.sheet_properties)
    target_sheet.page_margins = copy(source_sheet.page_margins)
    target_sheet.page_setup = copy(source_sheet.page_setup)
    target_sheet.print_options = copy(source_sheet.print_options)
    target_sheet.freeze_panes = source_sheet.freeze_panes
    target_sheet.sheet_view.zoomScale = source_sheet.sheet_view.zoomScale

    for row in source_sheet.iter_rows():
        for source_cell in row:
            target_cell = target_sheet.cell(source_cell.row, source_cell.column)
            target_cell.value = source_cell.value
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            target_cell.number_format = source_cell.number_format
            target_cell.font = copy(source_cell.font)
            target_cell.fill = copy(source_cell.fill)
            target_cell.border = copy(source_cell.border)
            target_cell.alignment = copy(source_cell.alignment)
            target_cell.protection = copy(source_cell.protection)

    for row_index, row_dimension in source_sheet.row_dimensions.items():
        target_dimension = target_sheet.row_dimensions[row_index]
        target_dimension.height = row_dimension.height
        target_dimension.hidden = row_dimension.hidden

    for col_key, col_dimension in source_sheet.column_dimensions.items():
        target_dimension = target_sheet.column_dimensions[col_key]
        target_dimension.width = col_dimension.width
        target_dimension.hidden = col_dimension.hidden
        target_dimension.bestFit = col_dimension.bestFit

    for merged_range in source_sheet.merged_cells.ranges:
        target_sheet.merge_cells(str(merged_range))

    return target_sheet


def clone_column_format(worksheet: Worksheet, source_col: int, target_col: int) -> None:
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    source_dimension = worksheet.column_dimensions[source_letter]
    target_dimension = worksheet.column_dimensions[target_letter]
    target_dimension.width = source_dimension.width
    target_dimension.hidden = source_dimension.hidden
    target_dimension.bestFit = source_dimension.bestFit

    for row in range(1, worksheet.max_row + 1):
        source_cell = worksheet.cell(row, source_col)
        target_cell = worksheet.cell(row, target_col)
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)


def build_sheet_day_header_value(current_day: date) -> float:
    return float(f"{current_day.day}.{current_day.month:02d}")


def pick_ovine_count(primary, fallback) -> float:
    return as_number(fallback) if as_number(fallback) > 0 else as_number(primary)


def as_number(value) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip().replace(",", ".")
    try:
        return float(text)
    except ValueError:
        return 0.0


def is_total_row(value) -> bool:
    return normalize_label(value).startswith("total")


def unique_output_path(output_dir: Path, filename: str) -> Path:
    candidate = output_dir / filename
    if not candidate.exists():
        return candidate

    stem = candidate.stem
    suffix = candidate.suffix
    counter = 2
    while True:
        numbered = output_dir / f"{stem} ({counter}){suffix}"
        if not numbered.exists():
            return numbered
        counter += 1


def enable_full_recalculation(workbook: Workbook) -> None:
    calculation = getattr(workbook, "calculation", None)
    if calculation is None:
        return
    calculation.calcMode = "auto"
    calculation.fullCalcOnLoad = True
    calculation.forceFullCalc = True


def _emit(
    callback: Callable[[ProgressEvent], None] | None,
    percent: float,
    message: str,
) -> None:
    if callback is not None:
        callback(ProgressEvent(percent=percent, message=message))


def _cell_text(value) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _append_warning(warnings: list[str] | None, message: str) -> None:
    if warnings is not None:
        warnings.append(message)


def _is_pdf_path(path: str | Path) -> bool:
    return Path(path).suffix.lower() == ".pdf"


def _load_pdfplumber():
    try:
        import pdfplumber
    except ImportError as exc:
        raise ProcessorError("Lipsește dependența pdfplumber pentru citirea PDF-urilor.") from exc
    return pdfplumber


def _extract_pdf_text_lines(pdf_path: str | Path) -> list[str]:
    pdfplumber = _load_pdfplumber()
    lines: list[str] = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            lines.extend(line.strip() for line in text.splitlines() if line.strip())
    return lines


def _extract_pdf_table_rows(pdf_path: str | Path) -> list[list[str]]:
    pdfplumber = _load_pdfplumber()
    rows: list[list[str]] = []
    table_settings_options = [
        {
            "vertical_strategy": "lines",
            "horizontal_strategy": "lines",
            "intersection_tolerance": 5,
            "snap_tolerance": 3,
            "join_tolerance": 3,
        },
        {
            "vertical_strategy": "text",
            "horizontal_strategy": "text",
            "snap_tolerance": 3,
            "join_tolerance": 3,
        },
    ]

    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            page_tables = []
            for table_settings in table_settings_options:
                page_tables = page.extract_tables(table_settings=table_settings) or []
                if page_tables:
                    break

            if page_tables:
                for table in page_tables:
                    for row in table:
                        normalized = [_normalize_pdf_cell(cell) for cell in row]
                        while normalized and not normalized[-1]:
                            normalized.pop()
                        if any(normalized):
                            rows.append(normalized)
                continue

            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""
            for line in text.splitlines():
                split_row = _split_pdf_text_line(line)
                if any(split_row):
                    rows.append(split_row)

    return rows


def _normalize_pdf_cell(value) -> str:
    text = _cell_text(value)
    text = text.replace("\r", "\n")
    text = re.sub(r"\s*\n\s*", " ", text)
    return re.sub(r"\s+", " ", text).strip()


def _split_pdf_text_line(line: str) -> list[str]:
    text = _normalize_pdf_cell(line)
    if not text:
        return []
    return re.split(r"\s{2,}", text)


def _pdf_cell(row: list[str], index: int | None) -> str:
    if index is None or index < 0 or index >= len(row):
        return ""
    return _normalize_pdf_cell(row[index])


def _parse_full_date_from_text(value) -> date | None:
    text = _cell_text(value)
    if not text:
        return None
    parsed = parse_full_date(text)
    if parsed is not None:
        return parsed
    match = FULL_DATE_SEARCH_RE.search(text)
    if match is None:
        return None
    try:
        return date(
            int(match.group("year")),
            int(match.group("month")),
            int(match.group("day")),
        )
    except ValueError:
        return None


def _parse_month_interval_from_text(value) -> TargetMonth | None:
    text = _cell_text(value)
    match = DATE_INTERVAL_RE.search(text)
    if match is None:
        return None
    month_number = int(match.group("month"))
    if month_number not in range(1, 13):
        return None
    return TargetMonth(
        year=int(match.group("year")),
        month_number=month_number,
        month_name=MONTH_NAMES[month_number - 1],
    )


def _detect_month_from_pdf(pdf_path: str | Path) -> TargetMonth | None:
    lines = _extract_pdf_text_lines(pdf_path)
    if not lines:
        raise ProcessorError("PDF-ul nu conține text extractibil. OCR nu este implementat în această versiune.")

    for line in lines:
        month = _parse_month_from_text(line) or _parse_month_interval_from_text(line)
        if month is not None:
            return month

    for line in lines:
        parsed_day = _parse_full_date_from_text(line)
        if parsed_day is not None:
            return TargetMonth(parsed_day.year, parsed_day.month, MONTH_NAMES[parsed_day.month - 1])
    return None


def _extract_transatori_entries(value) -> list[str]:
    text = _cell_text(value)
    if not text:
        return []
    entries: list[str] = []
    for match in TRANSATORI_RE.finditer(text):
        count = float(match.group("count").replace(",", "."))
        if count <= 30:
            entries.append(match.group(0))
    return entries


def _find_attendance_pdf_header(rows: list[list[str]]) -> tuple[int, dict[str, int | None]] | None:
    for row_index, row in enumerate(rows[:30]):
        normalized = [normalize_label(cell) for cell in row]
        columns = {
            "name": _find_pdf_header_column(normalized, "nume si prenume"),
            "co": _find_pdf_header_column(normalized, "co"),
            "n": _find_pdf_header_column(normalized, "n"),
            "mentiuni": _find_pdf_header_column(normalized, "mentiuni"),
        }
        if columns["name"] is not None and columns["co"] is not None and columns["mentiuni"] is not None:
            return row_index, columns
    return None


def _find_pdf_header_column(normalized_cells: Iterable[str], expected: str) -> int | None:
    for index, cell in enumerate(normalized_cells):
        if cell == expected:
            return index
    return None


def _looks_like_attendance_data_row(row: list[str]) -> bool:
    first_cell = _pdf_cell(row, 0)
    if first_cell.isdigit():
        return True
    last_cell = _normalize_pdf_cell(row[-1]) if row else ""
    return bool(_pdf_cell(row, 2) or _pdf_cell(row, 3) or last_cell)


def _detect_month_from_sheet(worksheet: Worksheet) -> TargetMonth | None:
    max_row = min(worksheet.max_row, 12)
    max_col = min(worksheet.max_column, 8)
    for row in range(1, max_row + 1):
        for col in range(1, max_col + 1):
            month = _parse_month_from_text(worksheet.cell(row, col).value)
            if month is not None:
                return month
    return None


def _find_attendance_header(worksheet: Worksheet) -> AttendanceHeader | None:
    for row in range(1, min(worksheet.max_row, 20) + 1):
        columns_by_label: dict[str, int] = {}
        for col in range(1, worksheet.max_column + 1):
            label = normalize_label(worksheet.cell(row, col).value)
            if label:
                columns_by_label[label] = col

        name_col = columns_by_label.get("nume si prenume")
        co_col = columns_by_label.get("co")
        n_col = columns_by_label.get("n")
        mentiuni_col = columns_by_label.get("mentiuni")
        if name_col and co_col and mentiuni_col:
            return AttendanceHeader(
                header_row=row,
                name_col=name_col,
                co_col=co_col,
                n_col=n_col,
                mentiuni_col=mentiuni_col,
            )
    return None


def _parse_month_from_text(value) -> TargetMonth | None:
    if value is None:
        return None
    match = SOURCE_MONTH_RE.search(normalize_label(value))
    if match is None:
        return None
    month_name = normalize_label(match.group("month"))
    year = int(match.group("year"))
    month_number = MONTH_NAME_TO_NUMBER[month_name]
    return TargetMonth(year=year, month_number=month_number, month_name=month_name)
