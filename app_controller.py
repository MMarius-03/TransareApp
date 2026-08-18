from __future__ import annotations

import os
import subprocess
from datetime import date
from pathlib import Path

from openpyxl import load_workbook
from PySide6.QtCore import QObject, QThread, Signal
from PySide6.QtWidgets import QFileDialog

from app_window import (
    APP_TITLE,
    TEMPLATE_MODE_PREDEFINED_LABEL,
    TEMPLATE_MODE_PREVIOUS_SHEET_LABEL,
    choose_missing_source_days_dialog,
    choose_mismatched_source_dates_dialog,
    confirm_warning_dialog,
    show_error_dialog,
    show_warning_dialog,
)
from processor import (
    MONTH_NAMES,
    ProcessingCancelled,
    ProcessorError,
    ProgressEvent,
    RunResult,
    TEMPLATE_MODE_PREDEFINED,
    TEMPLATE_MODE_PREVIOUS_SHEET,
    build_target_day_columns,
    detect_attendance_month_from_file,
    detect_source_month_from_file,
    filter_target_days_to_source,
    find_source_date_month_mismatches,
    normalize_label,
    parse_source_days,
    run_fill,
)

TEMPLATE_MODE_BY_LABEL = {
    TEMPLATE_MODE_PREDEFINED_LABEL: TEMPLATE_MODE_PREDEFINED,
    TEMPLATE_MODE_PREVIOUS_SHEET_LABEL: TEMPLATE_MODE_PREVIOUS_SHEET,
}


class RunWorker(QObject):
    progress = Signal(float, str)
    success = Signal(object)
    error = Signal(str)
    cancelled = Signal()
    finished = Signal()

    def __init__(
        self,
        source_path: str,
        target_path: str,
        template_mode: str,
        target_month_name: str,
        template_sheet_name: str | None,
        attendance_path: str | None,
        allow_attendance_month_mismatch: bool,
        include_missing_source_days: bool,
        treat_mismatched_source_dates_as_target: bool,
    ) -> None:
        super().__init__()
        self.source_path = source_path
        self.target_path = target_path
        self.template_mode = template_mode
        self.target_month_name = target_month_name
        self.template_sheet_name = template_sheet_name
        self.attendance_path = attendance_path
        self.allow_attendance_month_mismatch = allow_attendance_month_mismatch
        self.include_missing_source_days = include_missing_source_days
        self.treat_mismatched_source_dates_as_target = treat_mismatched_source_dates_as_target
        self._is_cancelled = False

    def cancel(self) -> None:
        self._is_cancelled = True

    def run(self) -> None:
        try:
            result = run_fill(
                source_path=self.source_path,
                target_path=self.target_path,
                template_mode=self.template_mode,
                target_month_name=self.target_month_name,
                template_sheet_name=self.template_sheet_name,
                attendance_path=self.attendance_path,
                allow_attendance_month_mismatch=self.allow_attendance_month_mismatch,
                include_missing_source_days=self.include_missing_source_days,
                treat_mismatched_source_dates_as_target=self.treat_mismatched_source_dates_as_target,
                progress_callback=self._on_progress,
                cancel_check=lambda: self._is_cancelled,
            )
        except ProcessingCancelled:
            self.cancelled.emit()
        except ProcessorError as exc:
            self.error.emit(str(exc))
        except Exception as exc:
            self.error.emit(f"Eroare neașteptată: {exc}")
        else:
            self.success.emit(result)
        finally:
            self.finished.emit()

    def _on_progress(self, event: ProgressEvent) -> None:
        self.progress.emit(event.percent, event.message)


class AppController(QObject):
    def __init__(self, window, base_dir: Path | None = None) -> None:
        super().__init__()
        self.window = window
        self.base_dir = (base_dir or Path.cwd()).resolve()
        self.last_output_file: Path | None = None
        self.detected_source_month_name: str | None = None
        self.detected_source_year: int | None = None
        self._thread: QThread | None = None
        self._worker: RunWorker | None = None
        self._connect_window()

    def initialize(self) -> None:
        self.window.set_busy(False, last_output_exists=False)
        self.window.append_log("Aplicația este gata.")

    def _connect_window(self) -> None:
        self.window.browse_source_requested.connect(self.browse_source_file)
        self.window.browse_target_requested.connect(self.browse_target_file)
        self.window.browse_attendance_requested.connect(self.browse_attendance_file)
        self.window.run_requested.connect(self.start_run)
        self.window.cancel_requested.connect(self.cancel_run)
        self.window.open_output_requested.connect(self.open_last_output)

    def browse_source_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self.window,
            "Alege situația transatori",
            "",
            "Fișiere Excel sau PDF (*.xlsx *.pdf);;Fișiere Excel (*.xlsx);;Fișiere PDF (*.pdf);;Toate fișierele (*)",
        )
        if path:
            self.window.set_source_path(path)
            self.window.append_log(f"Sursă selectată: {path}")
            try:
                detected_month = detect_source_month_from_file(path)
            except ProcessorError as exc:
                self.detected_source_month_name = None
                self.detected_source_year = None
                if hasattr(self.window, "set_source_month_hint"):
                    self.window.set_source_month_hint(None)
                self.window.append_log(f"Avertisment: {exc}")
                self.window.set_status("Fișierul sursă a fost selectat, dar luna nu a putut fi detectată.", "warning")
            else:
                self.detected_source_month_name = detected_month.month_name if detected_month is not None else None
                self.detected_source_year = detected_month.year if detected_month is not None else None
                if hasattr(self.window, "set_source_month_hint"):
                    self.window.set_source_month_hint(self.detected_source_month_name)
                if self.detected_source_month_name:
                    self.window.set_target_month(self.detected_source_month_name)
                    self.window.append_log(f"Luna detectată din sursă: {self.detected_source_month_name}")
                self.window.set_status("Fișierul sursă a fost selectat.", "info")
                self._refresh_template_selection()

    def browse_target_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self.window,
            "Alege workbook-ul de salarii",
            "",
            "Excel files (*.xlsx)",
        )
        if not path:
            return

        self.window.set_target_path(path)
        self.window.append_log(f"Workbook salarii selectat: {path}")
        try:
            sheet_names = self._load_sheet_names(path)
        except ProcessorError as exc:
            self._show_error(str(exc))
            return

        selected = self._choose_default_template_sheet(sheet_names)
        self.window.set_template_sheet_options(sheet_names, selected=selected)
        self.window.set_status("Foile workbook-ului au fost încărcate.", "success")

    def browse_attendance_file(self) -> None:
        path, _ = QFileDialog.getOpenFileName(
            self.window,
            "Alege fișierul Transatori+Detinuți",
            "",
            "Fișiere Excel sau PDF (*.xlsx *.pdf);;Fișiere Excel (*.xlsx);;Fișiere PDF (*.pdf);;Toate fișierele (*)",
        )
        if path:
            self.window.set_attendance_path(path)
            self.window.append_log(f"Transatori+Detinuți selectat: {path}")
            self.window.set_status("Fișierul Transatori+Detinuți a fost selectat.", "info")

    def _load_sheet_names(self, workbook_path: str) -> list[str]:
        workbook = None
        try:
            workbook = load_workbook(workbook_path, read_only=True, data_only=True)
            return list(workbook.sheetnames)
        except Exception as exc:
            raise ProcessorError("Nu am putut citi foile din workbook-ul de salarii.") from exc
        finally:
            if workbook is not None:
                workbook.close()

    def start_run(self) -> None:
        if self._thread is not None:
            self._show_warning("Procesarea este deja în curs.")
            return

        values = self.window.form_values()
        missing = [
            key
            for key in ("source_path", "target_path", "target_month_name", "template_mode_label")
            if not values.get(key)
        ]
        if missing:
            self._show_warning("Completează toate câmpurile înainte de rulare.")
            return

        template_mode = TEMPLATE_MODE_BY_LABEL.get(values["template_mode_label"])
        if template_mode is None:
            self._show_warning("Modul de template selectat nu este valid.")
            return

        if (
            self.detected_source_month_name
            and normalize_label(values["target_month_name"]) != self.detected_source_month_name
        ):
            self._show_warning(
                f"Luna detectată în sursă este {self.detected_source_month_name}, nu {values['target_month_name']}."
            )
            return

        parsed_source_days, source_target_year = self._load_source_days_for_target_month(
            values["source_path"],
            values["target_month_name"],
        )
        treat_mismatched_source_dates_as_target = False
        mismatched_source_dates = self._find_source_date_month_mismatches(
            parsed_source_days,
            source_target_year,
            values["target_month_name"],
        )
        if mismatched_source_dates:
            treat_mismatched_source_dates_as_target = self._choose_mismatched_source_dates(
                mismatched_source_dates
            )
            if treat_mismatched_source_dates_as_target is None:
                self.window.append_log(
                    "Procesarea a fost oprită deoarece există date din altă lună în fișierul sursă."
                )
                return
            if treat_mismatched_source_dates_as_target:
                formatted_mismatches = ", ".join(
                    f"{source_day.strftime('%d.%m.%Y')} → {target_day.strftime('%d.%m.%Y')}"
                    for source_day, target_day in mismatched_source_dates
                )
                self.window.append_log(
                    f"Date tratate ca luna aleasă numai în raport: {formatted_mismatches}."
                )

        include_missing_source_days = True
        missing_source_days = self._find_missing_source_days(
            parsed_source_days,
            source_target_year,
            values["target_month_name"],
            exclude_day_numbers={target_day.day for _source_day, target_day in mismatched_source_dates}
            if treat_mismatched_source_dates_as_target
            else set(),
        )
        if missing_source_days:
            include_missing_source_days = self._choose_missing_source_days(missing_source_days)
            if include_missing_source_days is None:
                self.window.append_log(
                    "Procesarea a fost oprită deoarece lipsesc zile din fișierul sursă."
                )
                return
            formatted_days = ", ".join(day.strftime("%d.%m.%Y") for day in missing_source_days)
            if include_missing_source_days:
                self.window.append_log(
                    f"Zile adăugate cu 0 în raport: {formatted_days}."
                )
            else:
                self.window.append_log(
                    f"Zile neincluse în raport: {formatted_days}."
                )

        if template_mode == TEMPLATE_MODE_PREVIOUS_SHEET and not values["template_sheet_name"]:
            self._show_warning("Alege un sheet template pentru modul de copiere din sheet anterior.")
            return

        if (
            template_mode == TEMPLATE_MODE_PREVIOUS_SHEET
            and normalize_label(values["template_sheet_name"]) == normalize_label(values["target_month_name"])
        ):
            self._show_warning("Sheet-ul template trebuie să fie luna anterioară sau un alt model, nu aceeași lună.")
            return

        attendance_path = values.get("attendance_path") or None
        allow_attendance_month_mismatch = False
        if attendance_path:
            if not Path(attendance_path).exists():
                self._show_warning("Fișierul Transatori+Detinuți nu există.")
                return
            try:
                attendance_month = detect_attendance_month_from_file(attendance_path)
            except ProcessorError as exc:
                self._show_warning(str(exc))
                return
            if attendance_month is not None:
                target_month_norm = normalize_label(values["target_month_name"])
                month_mismatch = attendance_month.month_name != target_month_norm
                year_mismatch = (
                    self.detected_source_year is not None
                    and attendance_month.year != self.detected_source_year
                )
                if month_mismatch or year_mismatch:
                    allow_attendance_month_mismatch = self._confirm_attendance_month_mismatch(
                        f"{attendance_month.month_name} {attendance_month.year}",
                        values["target_month_name"],
                    )
                    if not allow_attendance_month_mismatch:
                        self.window.append_log("Procesarea a fost oprită din cauza lunii diferite din Transatori+Detinuți.")
                        return

        self.window.set_busy(True, last_output_exists=self.last_output_file is not None)
        self.window.set_status("Pornesc procesarea workbook-ului...", "info")
        self.window.set_progress(3, "Inițializare")
        self.window.append_log("Pornesc generarea și completarea foii lunare.")

        self._thread = QThread(self)
        self._worker = RunWorker(
            source_path=values["source_path"],
            target_path=values["target_path"],
            template_mode=template_mode,
            target_month_name=values["target_month_name"],
            template_sheet_name=values["template_sheet_name"] or None,
            attendance_path=attendance_path,
            allow_attendance_month_mismatch=allow_attendance_month_mismatch,
            include_missing_source_days=include_missing_source_days,
            treat_mismatched_source_dates_as_target=treat_mismatched_source_dates_as_target,
        )
        self._worker.moveToThread(self._thread)
        self._thread.started.connect(self._worker.run)
        self._worker.progress.connect(self._handle_progress)
        self._worker.success.connect(self._handle_success)
        self._worker.error.connect(self._handle_error)
        self._worker.cancelled.connect(self._handle_cancelled)
        self._worker.finished.connect(self._cleanup_worker)
        self._thread.start()

    def cancel_run(self) -> None:
        if self._worker is None:
            return
        self._worker.cancel()
        self.window.set_status("Se anulează procesarea...", "warning")
        self.window.append_log("Anulare solicitată de utilizator.")

    def shutdown(self) -> None:
        """Cancel any running worker and wait for the thread to finish (used on close)."""
        thread = self._thread
        if thread is None:
            return
        if self._worker is not None:
            self._worker.cancel()
        thread.quit()
        thread.wait(10000)

    def _handle_progress(self, percent: float, message: str) -> None:
        self.window.set_progress(percent, message)
        self.window.set_status(message, "info")

    def _handle_success(self, result: RunResult) -> None:
        self.last_output_file = result.output_file
        self.window.set_open_output_enabled(True)
        self.window.set_progress(100, "Completare finalizată")
        self.window.set_status(
            f"Workbook actualizat: {result.output_file.name}",
            "success",
        )
        self.window.append_log(f"Workbook actualizat: {result.output_file}")
        self.window.append_log(
            "Foi create: "
            f"{result.created_sheet_name}, {result.test_sheet_name} | "
            f"template: {result.template_source_name or 'preset'}"
        )
        self.window.append_log(f"Zile mapate: {result.mapped_days}; blocuri actualizate: {result.updated_blocks}")
        attendance_summary = getattr(result, "attendance_summary", None)
        if attendance_summary is not None:
            self.window.append_log(
                "Transatori+Detinuți: "
                f"{attendance_summary.matched_employees} angajați potriviți, "
                f"{attendance_summary.approximate_matches} potriviri aproximative, "
                f"{attendance_summary.co_days_applied} zile CO aplicate, "
                f"{attendance_summary.n_days_applied} zile N aplicate, "
                f"{attendance_summary.cm_days_applied} zile CM aplicate, "
                f"{attendance_summary.mentiuni_days_colored} zile din mențiuni colorate, "
                f"{attendance_summary.mentiuni_copied} mențiuni copiate, "
                f"{len(attendance_summary.unmatched_names)} nepotriviți, "
                f"{len(attendance_summary.ambiguous_names)} ambigui."
            )
            for name in attendance_summary.unmatched_names:
                self.window.append_log(f"Transatori+Detinuți fără potrivire: {name}")
            for name in attendance_summary.ambiguous_names:
                self.window.append_log(f"Transatori+Detinuți ambiguu: {name}")
            for name in attendance_summary.added_employees:
                self.window.append_log(f"Angajat nou adăugat automat: {name}")
        for warning in result.warnings:
            self.window.append_log(f"Avertisment: {warning}")
        if result.differences:
            self.window.append_log(f"Diferențe detectate față de sheetul curent: {len(result.differences)}")

    def _handle_error(self, message: str) -> None:
        self.window.set_progress(0, "Eroare")
        self.window.set_status(message, "error")
        self.window.append_log(message)
        self._show_error(message)

    def _handle_cancelled(self) -> None:
        self.window.set_progress(0, "Anulat")
        self.window.set_status("Procesarea a fost anulată.", "warning")
        self.window.append_log("Procesarea a fost anulată.")

    def _cleanup_worker(self) -> None:
        thread = self._thread
        worker = self._worker
        self._thread = None
        self._worker = None
        if thread is not None:
            thread.quit()
            thread.wait(1500)
            thread.deleteLater()
        if worker is not None:
            worker.deleteLater()
        self.window.set_busy(False, last_output_exists=self.last_output_file is not None)

    def open_last_output(self) -> None:
        if self.last_output_file is None or not self.last_output_file.exists():
            self._show_warning("Nu există încă un fișier final disponibil.")
            return

        try:
            if os.name == "nt":
                os.startfile(str(self.last_output_file))
            else:
                subprocess.run(["xdg-open", str(self.last_output_file)], check=False)
        except Exception as exc:
            self._show_error(f"Nu am putut deschide workbook-ul actualizat: {exc}")

    def _show_warning(self, message: str) -> None:
        show_warning_dialog(self.window, message, APP_TITLE)

    def _show_error(self, message: str) -> None:
        show_error_dialog(self.window, message, APP_TITLE)

    def _confirm_attendance_month_mismatch(self, attendance_month: str, target_month_name: str) -> bool:
        return confirm_warning_dialog(
            self.window,
            "Luna din Transatori+Detinuți pare diferită.\n\n"
            f"Fișier Transatori+Detinuți: {attendance_month}\n"
            f"Luna țintă selectată: {target_month_name}\n\n"
            "Vrei să continui oricum?",
            APP_TITLE,
        )

    def _load_source_days_for_target_month(
        self,
        source_path: str,
        target_month_name: str,
    ) -> tuple[dict, int | None]:
        path = Path(source_path)
        if not path.exists():
            return {}, None
        try:
            parsed_days = parse_source_days(path)
        except ProcessorError:
            return {}, None

        month_name = normalize_label(target_month_name)
        month_number = MONTH_NAMES.index(month_name) + 1 if month_name in MONTH_NAMES else None
        if month_number is None:
            return parsed_days, None
        dates_in_target_month = [day for day in parsed_days if day.month == month_number]
        target_year = dates_in_target_month[0].year if dates_in_target_month else None
        return parsed_days, target_year

    def _find_source_date_month_mismatches(
        self,
        parsed_days: dict,
        target_year: int | None,
        target_month_name: str,
    ) -> list[tuple[date, date]]:
        month_name = normalize_label(target_month_name)
        if not parsed_days or target_year is None or month_name not in MONTH_NAMES:
            return []
        month_number = MONTH_NAMES.index(month_name) + 1
        return [
            (mismatch.source_date, mismatch.target_date)
            for mismatch in find_source_date_month_mismatches(
                parsed_days,
                target_year,
                month_number,
            )
        ]

    def _find_missing_source_days(
        self,
        parsed_days: dict,
        target_year: int | None,
        target_month_name: str,
        *,
        exclude_day_numbers: set[int],
    ) -> list[date]:
        """Return target-month working days that do not occur in the source file.

        This is a non-blocking preflight check. The worker remains responsible for
        reporting malformed or unreadable source files with the standard error.
        """
        if not parsed_days or target_year is None:
            return []

        month_name = normalize_label(target_month_name)
        month_number = MONTH_NAMES.index(month_name) + 1 if month_name in MONTH_NAMES else None
        if month_number is None:
            return []

        target_days = build_target_day_columns(target_year, month_number)
        _included_days, missing_days = filter_target_days_to_source(target_days, parsed_days)
        return [day for day in missing_days if day.day not in exclude_day_numbers]

    def _choose_mismatched_source_dates(
        self,
        mismatched_dates: list[tuple[date, date]],
    ) -> bool | None:
        formatted_mismatches = ", ".join(
            f"{source_day.strftime('%d.%m.%Y')} → {target_day.strftime('%d.%m.%Y')}"
            for source_day, target_day in mismatched_dates
        )
        return choose_mismatched_source_dates_dialog(
            self.window,
            "Fișierul Situație are date într-o lună diferită, iar zilele respective "
            "lipsesc din luna aleasă:\n"
            f"{formatted_mismatches}\n\n"
            "Vrei să le tratezi ca zile din luna aleasă? Corecția se aplică numai "
            "în raportul generat și nu modifică fișierul sursă.",
            APP_TITLE,
        )

    def _choose_missing_source_days(self, missing_days: list[date]) -> bool | None:
        formatted_days = ", ".join(day.strftime("%d.%m.%Y") for day in missing_days)
        return choose_missing_source_days_dialog(
            self.window,
            "Fișierul Situație nu conține date pentru:\n"
            f"{formatted_days}\n\n"
            "Vrei să le adaugi în raportul final cu valoarea 0 și culoarea galbenă?\n\n"
            "Poți și să continui fără aceste zile în raport sau să oprești procesarea.",
            APP_TITLE,
        )

    def _refresh_template_selection(self) -> None:
        values = self.window.form_values()
        target_path = values.get("target_path", "")
        if not target_path or not Path(target_path).exists():
            return
        try:
            sheet_names = self._load_sheet_names(target_path)
        except ProcessorError:
            return
        self.window.set_template_sheet_options(sheet_names, selected=self._choose_default_template_sheet(sheet_names))

    def _choose_default_template_sheet(self, sheet_names: list[str]) -> str:
        if not sheet_names:
            return ""

        target_month = normalize_label(self.window.form_values().get("target_month_name", ""))
        previous_month = self._previous_month_name(target_month)
        if previous_month:
            for name in sheet_names:
                if normalize_label(name) == previous_month:
                    return name

        for preferred in ("martie", *MONTH_NAMES):
            for name in sheet_names:
                if normalize_label(name) == preferred:
                    return name
        return sheet_names[0]

    def _previous_month_name(self, month_name: str) -> str | None:
        if month_name not in MONTH_NAMES:
            return None
        index = MONTH_NAMES.index(month_name)
        if index == 0:
            return MONTH_NAMES[-1]
        return MONTH_NAMES[index - 1]
